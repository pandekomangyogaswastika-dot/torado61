"""Bulk Excel Import Service

Supports template generation, parsing, validation, and import for master data:
- Items (inventory items)
- Vendors
- Employees
- Chart of Accounts (COA)
- Customers

Features:
- Download Excel template with headers
- Parse uploaded Excel file
- Validate data with detailed error reporting per row
- Upsert (create/update) based on unique key
- Return summary (created/updated/skipped/errors)
"""
import io
import logging
import uuid
from datetime import datetime, timezone

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.db import get_db
from core.exceptions import ValidationError

logger = logging.getLogger("aurora.excel_import")


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


# ============================================================================
# ENTITY CONFIGURATIONS
# ============================================================================

ENTITY_CONFIGS = {
    "items": {
        "collection": "items",
        "label": "Inventory Items",
        "unique_key": "code",  # upsert based on code
        "template_columns": [
            ("code", "Item Code*", "SKU001"),
            ("name", "Item Name*", "Minyak Goreng 1L"),
            ("category", "Category*", "Kitchen / Bar / Beverage"),
            ("unit", "Unit*", "pcs / kg / liter"),
            ("min_stock", "Min Stock", "10"),
            ("max_stock", "Max Stock", "100"),
            ("reorder_point", "Reorder Point", "20"),
            ("cost_method", "Cost Method", "avg / fifo / std"),
            ("std_cost", "Standard Cost", "15000"),
        ],
        "required_fields": ["code", "name", "category", "unit"],
    },
    "vendors": {
        "collection": "vendors",
        "label": "Vendors / Suppliers",
        "unique_key": "code",
        "template_columns": [
            ("code", "Vendor Code*", "VEN001"),
            ("name", "Vendor Name*", "PT Supplier ABC"),
            ("contact_person", "Contact Person", "Budi Santoso"),
            ("phone", "Phone", "021-12345678"),
            ("email", "Email", "vendor@example.com"),
            ("address", "Address", "Jl. Sudirman No. 123"),
            ("city", "City", "Jakarta"),
            ("payment_terms_days", "Payment Terms (days)", "30"),
            ("bank_name", "Bank Name", "BCA"),
            ("bank_account", "Bank Account Number", "1234567890"),
        ],
        "required_fields": ["code", "name"],
    },
    "employees": {
        "collection": "employees",
        "label": "Employees",
        "unique_key": "email",
        "template_columns": [
            ("email", "Email*", "employee@torado.id"),
            ("name", "Full Name*", "John Doe"),
            ("nik", "NIK (Employee ID)", "EMP001"),
            ("phone", "Phone", "08123456789"),
            ("position", "Position", "Cashier / Chef / Manager"),
            ("department", "Department", "Outlet / Kitchen / Finance"),
            ("hire_date", "Hire Date", "2026-01-15"),
            ("salary", "Base Salary", "5000000"),
            ("bank_name", "Bank Name", "BCA"),
            ("bank_account", "Bank Account Number", "9876543210"),
        ],
        "required_fields": ["email", "name"],
    },
    "coa": {
        "collection": "coa",
        "label": "Chart of Accounts",
        "unique_key": "account_code",
        "template_columns": [
            ("account_code", "Account Code*", "1-1010"),
            ("account_name", "Account Name*", "Kas Bank BCA"),
            ("account_type", "Account Type*", "asset / liability / equity / revenue / expense"),
            ("subtype", "Subtype", "cash / ar / inventory / ap / sales / cogs / opex"),
            ("parent_code", "Parent Account Code", "1-1000"),
            ("is_header", "Is Header", "FALSE"),
            ("normal_balance", "Normal Balance*", "debit / credit"),
        ],
        "required_fields": ["account_code", "account_name", "account_type", "normal_balance"],
    },
    "customers": {
        "collection": "ar_customers",
        "label": "AR Customers",
        "unique_key": "code",
        "template_columns": [
            ("code", "Customer Code*", "CUST001"),
            ("name", "Customer Name*", "PT Client XYZ"),
            ("contact_person", "Contact Person", "Jane Doe"),
            ("phone", "Phone", "021-98765432"),
            ("email", "Email", "customer@example.com"),
            ("address", "Address", "Jl. Thamrin No. 456"),
            ("city", "City", "Jakarta"),
            ("credit_limit", "Credit Limit", "50000000"),
            ("payment_terms_days", "Payment Terms (days)", "14"),
        ],
        "required_fields": ["code", "name"],
    },
}


# ============================================================================
# TEMPLATE GENERATION
# ============================================================================

def generate_template(entity_type: str) -> bytes:
    """Generate Excel template with headers and example data."""
    if entity_type not in ENTITY_CONFIGS:
        raise ValidationError(f"Entity type tidak didukung: {entity_type}")

    config = ENTITY_CONFIGS[entity_type]
    wb = Workbook()
    ws = wb.active
    # Sanitize sheet title (Excel doesn't allow: / \ ? * [ ] :)
    sheet_title = config["label"].replace("/", "-").replace("\\", "-")[:31]
    ws.title = sheet_title

    # Header row styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Write headers
    for col_idx, (field_key, header_label, example) in enumerate(config["template_columns"], start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header_label
        cell.fill = header_fill
        cell.font = header_font

    # Write example row
    for col_idx, (field_key, header_label, example) in enumerate(config["template_columns"], start=1):
        ws.cell(row=2, column=col_idx, value=example)

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================================
# PARSING & VALIDATION
# ============================================================================

def parse_excel(entity_type: str, file_bytes: bytes) -> dict:
    """Parse Excel file and validate data.
    
    Returns:
    {
        "entity_type": str,
        "total_rows": int,
        "valid_rows": list[dict],  # rows that passed validation
        "invalid_rows": list[dict], # rows with errors: {row_num, data, errors}
        "summary": {"valid": int, "invalid": int}
    }
    """
    if entity_type not in ENTITY_CONFIGS:
        raise ValidationError(f"Entity type tidak didukung: {entity_type}")

    config = ENTITY_CONFIGS[entity_type]

    MAX_ROWS = 1000  # Sprint D guardrail

    try:
        # Read Excel (skip example row 2, start from row 3)
        df = pd.read_excel(io.BytesIO(file_bytes), header=0)
        df = df.iloc[1:]  # Skip example row
        df = df.dropna(how="all")  # Drop empty rows
    except Exception as e:
        raise ValidationError(f"Gagal membaca file Excel: {str(e)}")

    # Guardrail: cap at MAX_ROWS
    if len(df) > MAX_ROWS:
        logger.warning(f"File has {len(df)} rows, capping at {MAX_ROWS}")
        df = df.iloc[:MAX_ROWS]

    # Map header labels to field keys
    column_mapping = {}
    for field_key, header_label, _ in config["template_columns"]:
        # Find column by matching header (case-insensitive, strip whitespace and *)
        clean_label = header_label.replace("*", "").strip()
        for col in df.columns:
            if clean_label.lower() in str(col).lower():
                column_mapping[col] = field_key
                break

    valid_rows = []
    invalid_rows = []

    for idx, row in df.iterrows():
        row_num = idx + 3  # Excel row number (header=1, example=2, data starts=3)
        row_data = {}
        errors = []

        # Map columns to fields
        for col, field_key in column_mapping.items():
            value = row[col]
            if pd.isna(value):
                value = None
            else:
                value = str(value).strip()
            row_data[field_key] = value

        # Validate required fields
        for required_field in config["required_fields"]:
            if not row_data.get(required_field):
                errors.append(f"{required_field} is required")

        # Entity-specific validations
        validation_errors = _validate_entity_data(entity_type, row_data)
        errors.extend(validation_errors)

        if errors:
            invalid_rows.append({
                "row_num": row_num,
                "data": row_data,
                "errors": errors,
            })
        else:
            valid_rows.append(row_data)

    return {
        "entity_type": entity_type,
        "total_rows": len(df),
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "summary": {
            "valid": len(valid_rows),
            "invalid": len(invalid_rows),
        },
    }


def _validate_entity_data(entity_type: str, data: dict) -> list[str]:
    """Entity-specific validation rules."""
    errors = []

    if entity_type == "items":
        if data.get("category") and data["category"] not in ["Kitchen", "Bar", "Beverage", "Cleaning", "Packaging"]:
            errors.append("category must be one of: Kitchen, Bar, Beverage, Cleaning, Packaging")
        if data.get("unit") and data["unit"] not in ["pcs", "kg", "liter", "box", "pack"]:
            errors.append("unit must be one of: pcs, kg, liter, box, pack")
        if data.get("cost_method") and data["cost_method"] not in ["avg", "fifo", "std"]:
            errors.append("cost_method must be one of: avg, fifo, std")

    elif entity_type == "coa":
        if data.get("account_type") and data["account_type"] not in ["asset", "liability", "equity", "revenue", "expense"]:
            errors.append("account_type must be one of: asset, liability, equity, revenue, expense")
        if data.get("normal_balance") and data["normal_balance"] not in ["debit", "credit"]:
            errors.append("normal_balance must be one of: debit, credit")

    elif entity_type == "employees":
        email = data.get("email")
        if email and "@" not in email:
            errors.append("email must be valid email format")

    return errors


# ============================================================================
# IMPORT (UPSERT)
# ============================================================================

async def commit_import(entity_type: str, valid_rows: list[dict], user_id: str) -> dict:
    """Import valid rows to database (upsert based on unique key).
    
    Returns:
    {
        "created": int,
        "updated": int,
        "skipped": int,
        "errors": list[dict],  # {data, error}
    }
    """
    if entity_type not in ENTITY_CONFIGS:
        raise ValidationError(f"Entity type tidak didukung: {entity_type}")

    config = ENTITY_CONFIGS[entity_type]
    collection_name = config["collection"]
    unique_key = config["unique_key"]

    db = get_db()
    collection = db[collection_name]

    created = 0
    updated = 0
    skipped = 0
    errors = []

    for row_data in valid_rows:
        try:
            unique_value = row_data.get(unique_key)
            if not unique_value:
                errors.append({"data": row_data, "error": f"{unique_key} is required for upsert"})
                skipped += 1
                continue

            # Check if exists
            existing = await collection.find_one({unique_key: unique_value})

            # Prepare document
            doc = _prepare_document(entity_type, row_data, user_id, is_update=bool(existing))

            if existing:
                # Update
                await collection.update_one(
                    {unique_key: unique_value},
                    {"$set": doc}
                )
                updated += 1
                logger.info(f"Updated {entity_type}: {unique_key}={unique_value}")
            else:
                # Create
                await collection.insert_one(doc)
                created += 1
                logger.info(f"Created {entity_type}: {unique_key}={unique_value}")

        except Exception as e:
            logger.error(f"Error importing row {row_data}: {str(e)}")
            errors.append({"data": row_data, "error": str(e)})
            skipped += 1

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }


def _prepare_document(entity_type: str, row_data: dict, user_id: str, is_update: bool) -> dict:
    """Prepare document for MongoDB insertion/update."""
    doc = {}

    if entity_type == "items":
        if not is_update:
            doc["id"] = str(uuid.uuid4())
        doc["code"] = row_data.get("code")
        doc["name"] = row_data.get("name")
        doc["category"] = row_data.get("category")
        doc["unit"] = row_data.get("unit")
        doc["min_stock"] = float(row_data.get("min_stock") or 0)
        doc["max_stock"] = float(row_data.get("max_stock") or 0)
        doc["reorder_point"] = float(row_data.get("reorder_point") or 0)
        doc["cost_method"] = row_data.get("cost_method") or "avg"
        doc["std_cost"] = float(row_data.get("std_cost") or 0)
        if not is_update:
            doc["created_at"] = _now()
            doc["created_by"] = user_id
        doc["updated_at"] = _now()
        doc["updated_by"] = user_id

    elif entity_type == "vendors":
        if not is_update:
            doc["id"] = str(uuid.uuid4())
        doc["code"] = row_data.get("code")
        doc["name"] = row_data.get("name")
        doc["contact_person"] = row_data.get("contact_person")
        doc["phone"] = row_data.get("phone")
        doc["email"] = row_data.get("email")
        doc["address"] = row_data.get("address")
        doc["city"] = row_data.get("city")
        doc["payment_terms_days"] = int(row_data.get("payment_terms_days") or 30)
        doc["bank_name"] = row_data.get("bank_name")
        doc["bank_account"] = row_data.get("bank_account")
        if not is_update:
            doc["created_at"] = _now()
        doc["updated_at"] = _now()

    elif entity_type == "employees":
        if not is_update:
            doc["id"] = str(uuid.uuid4())
        doc["email"] = row_data.get("email")
        doc["name"] = row_data.get("name")
        doc["nik"] = row_data.get("nik")
        doc["phone"] = row_data.get("phone")
        doc["position"] = row_data.get("position")
        doc["department"] = row_data.get("department")
        doc["hire_date"] = row_data.get("hire_date")
        doc["salary"] = float(row_data.get("salary") or 0)
        doc["bank_name"] = row_data.get("bank_name")
        doc["bank_account"] = row_data.get("bank_account")
        if not is_update:
            doc["created_at"] = _now()
        doc["updated_at"] = _now()

    elif entity_type == "coa":
        if not is_update:
            doc["id"] = str(uuid.uuid4())
        doc["account_code"] = row_data.get("account_code")
        doc["account_name"] = row_data.get("account_name")
        doc["account_type"] = row_data.get("account_type")
        doc["subtype"] = row_data.get("subtype")
        doc["parent_code"] = row_data.get("parent_code")
        doc["is_header"] = row_data.get("is_header", "").lower() in ["true", "yes", "1"]
        doc["normal_balance"] = row_data.get("normal_balance")
        if not is_update:
            doc["created_at"] = _now()
        doc["updated_at"] = _now()

    elif entity_type == "customers":
        if not is_update:
            doc["id"] = str(uuid.uuid4())
        doc["code"] = row_data.get("code")
        doc["name"] = row_data.get("name")
        doc["contact_person"] = row_data.get("contact_person")
        doc["phone"] = row_data.get("phone")
        doc["email"] = row_data.get("email")
        doc["address"] = row_data.get("address")
        doc["city"] = row_data.get("city")
        doc["credit_limit"] = float(row_data.get("credit_limit") or 0)
        doc["payment_terms_days"] = int(row_data.get("payment_terms_days") or 14)
        if not is_update:
            doc["created_at"] = _now()
        doc["updated_at"] = _now()

    return doc
