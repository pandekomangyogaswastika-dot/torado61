"""reports_excel_sales_service — Excel export functions."""
from __future__ import annotations
import logging
from datetime import datetime, timezone
from typing import Optional
from core.db import get_db
from core.exceptions import ValidationError
logger = logging.getLogger("aurora.reports")

def _now():
    return datetime.now(timezone.utc).isoformat()

def _parse_date(s, *, fallback_today=False):
    if not s:
        return datetime.now(timezone.utc) if fallback_today else None
    try:
        return datetime.strptime(s, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError as e:
        raise ValidationError(f"Invalid date format: {s}") from e

async def generate_daily_sales_excel(
    *,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    outlet_ids: Optional[list[str]] = None,
    brand_ids: Optional[list[str]] = None,
):
    """Generate Daily Sales Summary Excel report.
    
    Columns: Date, Outlet, Brand, Grand Total, Transaction Count, Status
    Features: Subtotals per outlet, Grand total row, Styling
    """
    from services.excel_export_service import (
        create_workbook,
        add_report_header,
        write_table_headers,
        write_data_row,
        write_total_row,
        autosize_columns,
        freeze_header_row,
        apply_currency_format,
        format_date,
    )
    
    db = get_db()
    
    # Build query
    match_filter: dict = {"deleted_at": None, "status": "validated"}
    if date_from or date_to:
        match_filter["sales_date"] = {}
        if date_from:
            match_filter["sales_date"]["$gte"] = date_from
        if date_to:
            match_filter["sales_date"]["$lte"] = date_to
    if outlet_ids:
        match_filter["outlet_id"] = {"$in": outlet_ids}
    if brand_ids:
        match_filter["brand_id"] = {"$in": brand_ids}
    
    # Load data
    sales_docs = []
    async for doc in db.daily_sales.find(match_filter).sort([("sales_date", 1), ("outlet_id", 1)]):
        sales_docs.append(doc)
    
    # Load lookups
    outlet_map = {}
    async for o in db.outlets.find({"deleted_at": None}):
        outlet_map[o["id"]] = o.get("name", o["id"])
    
    brand_map = {}
    async for b in db.brands.find({"deleted_at": None}):
        brand_map[b["id"]] = b.get("name", b["id"])
    
    # Create workbook
    wb = create_workbook("Daily Sales")
    ws = wb.active
    
    # Add header
    subtitle = f"Period: {format_date(date_from) if date_from else 'All'} - {format_date(date_to) if date_to else 'All'}"
    current_row = add_report_header(
        ws,
        title="Daily Sales Summary Report",
        subtitle=subtitle,
    )
    
    # Table headers
    headers = ["Date", "Outlet", "Brand", "Grand Total", "Transaction Count", "Status"]
    current_row = write_table_headers(ws, headers, current_row)
    
    # Data rows
    grand_total = 0.0
    grand_trx_count = 0
    
    for doc in sales_docs:
        outlet_name = outlet_map.get(doc.get("outlet_id"), doc.get("outlet_id") or "")
        brand_name = brand_map.get(doc.get("brand_id"), doc.get("brand_id") or "")
        sales_total = float(doc.get("grand_total", 0) or 0)
        trx_count = int(doc.get("transaction_count", 0) or 0)
        
        row_data = [
            format_date(doc.get("sales_date")),
            outlet_name,
            brand_name,
            sales_total,
            trx_count,
            doc.get("status", ""),
        ]
        write_data_row(ws, row_data, current_row, number_cols=[4, 5])
        
        grand_total += sales_total
        grand_trx_count += trx_count
        current_row += 1
    
    # Total row
    if sales_docs:
        total_row = ["", "", "TOTAL", grand_total, grand_trx_count, ""]
        write_total_row(ws, total_row, current_row, number_cols=[4, 5])
    
    # Styling
    apply_currency_format(ws, 4, current_row - len(sales_docs), current_row, currency="Rp")
    autosize_columns(ws)
    freeze_header_row(ws, header_row=current_row - len(sales_docs) - 1)
    
    return wb


async def generate_outlet_performance_excel(
    *,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    outlet_ids: Optional[list[str]] = None,
):
    """Generate Outlet Performance Excel report.
    
    Columns: Outlet, Total Sales, Avg Daily Sales, Transaction Count, Days Active
    Features: Bar chart comparison, Conditional formatting
    """
    from services.excel_export_service import (
        create_workbook,
        add_report_header,
        write_table_headers,
        write_data_row,
        write_total_row,
        autosize_columns,
        freeze_header_row,
        apply_currency_format,
        format_date,
        add_bar_chart,
    )
    
    db = get_db()
    
    # Build query
    match_filter: dict = {"deleted_at": None, "status": "validated"}
    if date_from or date_to:
        match_filter["sales_date"] = {}
        if date_from:
            match_filter["sales_date"]["$gte"] = date_from
        if date_to:
            match_filter["sales_date"]["$lte"] = date_to
    if outlet_ids:
        match_filter["outlet_id"] = {"$in": outlet_ids}
    
    # Aggregate per outlet
    pipeline = [
        {"$match": match_filter},
        {
            "$group": {
                "_id": "$outlet_id",
                "total_sales": {"$sum": "$grand_total"},
                "trx_count": {"$sum": "$transaction_count"},
                "days_count": {"$sum": 1},
            }
        },
        {"$sort": {"total_sales": -1}},
    ]
    
    results = []
    async for doc in db.daily_sales.aggregate(pipeline):
        results.append(doc)
    
    # Load outlet names
    outlet_map = {}
    async for o in db.outlets.find({"deleted_at": None}):
        outlet_map[o["id"]] = o.get("name", o["id"])
    
    # Create workbook
    wb = create_workbook("Outlet Performance")
    ws = wb.active
    
    # Add header
    subtitle = f"Period: {format_date(date_from) if date_from else 'All'} - {format_date(date_to) if date_to else 'All'}"
    current_row = add_report_header(
        ws,
        title="Outlet Performance Report",
        subtitle=subtitle,
    )
    
    # Table headers
    headers = ["Outlet", "Total Sales", "Days Active", "Avg Daily Sales", "Transaction Count"]
    current_row = write_table_headers(ws, headers, current_row)
    data_start_row = current_row
    
    # Data rows
    grand_total = 0.0
    grand_trx = 0
    
    for doc in results:
        outlet_name = outlet_map.get(doc["_id"], doc["_id"] or "Unknown")
        total_sales = float(doc.get("total_sales", 0) or 0)
        days = int(doc.get("days_count", 0) or 0)
        avg_daily = total_sales / days if days > 0 else 0
        trx_count = int(doc.get("trx_count", 0) or 0)
        
        row_data = [
            outlet_name,
            total_sales,
            days,
            avg_daily,
            trx_count,
        ]
        write_data_row(ws, row_data, current_row, number_cols=[2, 3, 4, 5])
        
        grand_total += total_sales
        grand_trx += trx_count
        current_row += 1
    
    # Total row
    if results:
        total_days = sum(int(r.get("days_count", 0) or 0) for r in results)
        total_row = ["TOTAL", grand_total, total_days, grand_total / total_days if total_days > 0 else 0, grand_trx]
        write_total_row(ws, total_row, current_row, number_cols=[2, 3, 4, 5])
    
    # Styling
    apply_currency_format(ws, 2, data_start_row, current_row, currency="Rp")
    apply_currency_format(ws, 4, data_start_row, current_row, currency="Rp")
    autosize_columns(ws)
    freeze_header_row(ws, header_row=data_start_row - 1)
    
    # Add bar chart (Total Sales)
    if results and len(results) > 1:
        data_range = f"B{data_start_row}:B{current_row - 1}"
        cats_range = f"A{data_start_row}:A{current_row - 1}"
        add_bar_chart(ws, "Total Sales by Outlet", data_range, cats_range, position="G2")
    
    return wb


async def generate_fdo_history_excel(
    *,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    outlet_ids: Optional[list[str]] = None,
    status: Optional[str] = None,
):
    """Generate FDO History Excel report.
    
    Columns: Doc No, Date, Outlet, Items Summary, Status, Approved By
    Features: Status color coding
    """
    from services.excel_export_service import (
        create_workbook,
        add_report_header,
        write_table_headers,
        write_data_row,
        autosize_columns,
        freeze_header_row,
        format_date,
        format_datetime,
    )
    from openpyxl.styles import PatternFill
    
    db = get_db()
    
    # Build query — canonical store is `kdo_bdo_orders` (date field: req_date, type field: kind)
    match_filter: dict = {"deleted_at": None}
    if date_from or date_to:
        match_filter["req_date"] = {}
        if date_from:
            match_filter["req_date"]["$gte"] = date_from
        if date_to:
            match_filter["req_date"]["$lte"] = date_to
    if outlet_ids:
        match_filter["outlet_id"] = {"$in": outlet_ids}
    if status:
        match_filter["status"] = status
    
    # Load data
    fdo_docs = []
    async for doc in db.kdo_bdo_orders.find(match_filter).sort([("req_date", -1)]):
        fdo_docs.append(doc)
    
    # Load lookups
    outlet_map = {}
    async for o in db.outlets.find({"deleted_at": None}):
        outlet_map[o["id"]] = o.get("name", o["id"])
    
    user_map = {}
    async for u in db.users.find({"deleted_at": None}):
        user_map[u["id"]] = u.get("name", u.get("email", u["id"]))
    
    # Create workbook
    wb = create_workbook("FDO History")
    ws = wb.active
    
    # Add header
    subtitle = f"Period: {format_date(date_from) if date_from else 'All'} - {format_date(date_to) if date_to else 'All'}"
    current_row = add_report_header(
        ws,
        title="FDO History Report",
        subtitle=subtitle,
    )
    
    # Table headers
    headers = ["Doc No", "Request Date", "Outlet", "Items Count", "Status", "Approved By", "Approved At"]
    current_row = write_table_headers(ws, headers, current_row)
    data_start_row = current_row
    
    # Data rows
    for doc in fdo_docs:
        outlet_name = outlet_map.get(doc.get("outlet_id"), doc.get("outlet_id") or "")
        items_count = len(doc.get("lines", []))
        status_val = doc.get("status", "draft")
        approved_by_name = user_map.get(doc.get("approved_by"), "") if doc.get("approved_by") else ""
        approved_at = format_datetime(doc.get("approved_at")) if doc.get("approved_at") else ""
        
        row_data = [
            doc.get("doc_no", ""),
            format_date(doc.get("req_date")),
            outlet_name,
            items_count,
            status_val,
            approved_by_name,
            approved_at,
        ]
        write_data_row(ws, row_data, current_row, number_cols=[4])
        
        # Color code status cell
        status_cell = ws.cell(row=current_row, column=5)
        if status_val == "approved":
            status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        elif status_val == "rejected":
            status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        elif status_val == "pending":
            status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        current_row += 1
    
    # Styling
    autosize_columns(ws)
    freeze_header_row(ws, header_row=data_start_row - 1)
    
    return wb



# ============================================================
# 8. INVENTORY REPORTS (Phase 4.2)
# ============================================================

