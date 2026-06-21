"""Market List Service — quarterly reference price management.

Market List = harga acuan kuartalan (bukan harga real vendor).
Di-update manual oleh admin/procurement setiap awal kuartal.
Digunakan sebagai referensi/benchmark di KDO/BDO/FDO.
"""
import io
import logging
from datetime import datetime, timezone
from typing import Optional

from core.db import get_db, serialize
from core.exceptions import NotFoundError, ValidationError
from models.market_list import make_quarter, make_ml_price

logger = logging.getLogger("aurora.market_list")


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


# =================== QUARTERS ===================

async def list_quarters(year: Optional[int] = None) -> list[dict]:
    db = get_db()
    q: dict = {"deleted_at": None} if await _has_deleted_field() else {}
    if year:
        q["year"] = year
    docs = await db.market_list_quarters.find(q).sort([
        ("year", -1), ("quarter", -1)
    ]).to_list(100)
    return [serialize(d) for d in docs]


async def _has_deleted_field() -> bool:
    return False  # Market list quarters don't use soft delete


async def get_active_quarter() -> Optional[dict]:
    """Get the currently active quarter. Returns most recently activated one."""
    db = get_db()
    doc = await db.market_list_quarters.find_one(
        {"status": "active"},
        sort=[("year", -1), ("quarter", -1)]
    )
    return serialize(doc) if doc else None


async def get_quarter(quarter_id: str) -> dict:
    db = get_db()
    doc = await db.market_list_quarters.find_one({"id": quarter_id})
    if not doc:
        raise NotFoundError("Quarter tidak ditemukan")
    return serialize(doc)


async def create_quarter(year: int, quarter: int, *, user: dict) -> dict:
    """Create a new market list quarter."""
    db = get_db()
    if quarter not in (1, 2, 3, 4):
        raise ValidationError("Quarter harus 1-4")
    # Check if already exists
    existing = await db.market_list_quarters.find_one({"year": year, "quarter": quarter})
    if existing:
        raise ValidationError(f"Q{quarter}-{year} sudah ada")
    doc = make_quarter(year, quarter, created_by=user["id"])
    await db.market_list_quarters.insert_one(doc)
    logger.info(f"Created quarter {doc['label']} by {user['id']}")
    return serialize(doc)


async def activate_quarter(quarter_id: str, *, user: dict) -> dict:
    """Activate a quarter (only one active at a time)."""
    db = get_db()
    doc = await db.market_list_quarters.find_one({"id": quarter_id})
    if not doc:
        raise NotFoundError("Quarter tidak ditemukan")
    # Close current active quarter
    await db.market_list_quarters.update_many(
        {"status": "active"},
        {"$set": {"status": "closed", "updated_at": _now()}}
    )
    # Activate this one
    await db.market_list_quarters.update_one(
        {"id": quarter_id},
        {"$set": {"status": "active", "updated_at": _now()}}
    )
    logger.info(f"Activated quarter {doc['label']} by {user['id']}")
    fresh = await db.market_list_quarters.find_one({"id": quarter_id})
    return serialize(fresh)


# =================== PRICES ===================

async def set_ref_price(
    quarter_id: str,
    item_id: str,
    unit: str,
    ref_price: float,
    notes: Optional[str] = None,
    *,
    user: dict,
) -> dict:
    """Set/update reference price for an item in a quarter."""
    db = get_db()
    quarter = await db.market_list_quarters.find_one({"id": quarter_id})
    if not quarter:
        raise NotFoundError("Quarter tidak ditemukan")
    # AUDIT FIX: Validate ref_price > 0 (tidak boleh 0 atau negatif)
    if ref_price <= 0:
        raise ValidationError("Harga referensi harus lebih besar dari 0", field="ref_price")
    item = await db.items.find_one({"id": item_id, "deleted_at": None})
    if not item:
        raise NotFoundError("Item tidak ditemukan")

    # Get previous quarter price for variance calculation
    prev_price_doc = await _get_previous_quarter_price(item_id, unit, quarter)
    previous_ref_price = prev_price_doc.get("ref_price") if prev_price_doc else None

    now = _now()
    existing = await db.market_list_prices.find_one({
        "quarter_id": quarter_id, "item_id": item_id, "unit": unit
    })
    if existing:
        variance_pct = None
        if previous_ref_price and previous_ref_price > 0:
            variance_pct = round(((ref_price - previous_ref_price) / previous_ref_price) * 100, 2)
        await db.market_list_prices.update_one(
            {"id": existing["id"]},
            {"$set": {
                "ref_price": ref_price,
                "previous_ref_price": previous_ref_price,
                "variance_pct": variance_pct,
                "notes": notes,
                "updated_at": now,
                "updated_by": user["id"],
            }}
        )
        fresh = await db.market_list_prices.find_one({"id": existing["id"]})
        return serialize(fresh)
    else:
        doc = make_ml_price(
            quarter_id=quarter_id,
            quarter_label=quarter["label"],
            item_id=item_id,
            unit=unit,
            ref_price=ref_price,
            previous_ref_price=previous_ref_price,
            notes=notes,
            created_by=user["id"],
        )
        await db.market_list_prices.insert_one(doc)
        return serialize(doc)


async def _get_previous_quarter_price(item_id: str, unit: str, current_quarter: dict) -> Optional[dict]:
    """Get price from the previous quarter for variance calculation."""
    db = get_db()
    year = current_quarter["year"]
    quarter = current_quarter["quarter"]
    # Find previous quarter (same year or previous year)
    if quarter == 1:
        prev_year, prev_q = year - 1, 4
    else:
        prev_year, prev_q = year, quarter - 1
    prev_quarter = await db.market_list_quarters.find_one({"year": prev_year, "quarter": prev_q})
    if not prev_quarter:
        return None
    return await db.market_list_prices.find_one({
        "quarter_id": prev_quarter["id"], "item_id": item_id, "unit": unit
    })


async def get_ref_price(
    item_id: str,
    *,
    quarter_id: Optional[str] = None,
    unit: Optional[str] = None,
) -> Optional[dict]:
    """Get reference price for item in given quarter (or active quarter if None)."""
    db = get_db()
    if not quarter_id:
        active = await get_active_quarter()
        if not active:
            return None
        quarter_id = active["id"]
    q: dict = {"quarter_id": quarter_id, "item_id": item_id}
    if unit:
        q["unit"] = unit
    doc = await db.market_list_prices.find_one(q, sort=[("created_at", -1)])
    return serialize(doc) if doc else None


async def get_ref_prices_bulk(
    item_ids: list[str],
    *,
    quarter_id: Optional[str] = None,
) -> dict[str, dict]:
    """Get reference prices for multiple items. Returns {item_id: price_doc}."""
    db = get_db()
    if not quarter_id:
        active = await get_active_quarter()
        quarter_id = active["id"] if active else None
    if not quarter_id:
        return {}
    docs = await db.market_list_prices.find({
        "quarter_id": quarter_id, "item_id": {"$in": item_ids}
    }).to_list(500)
    result = {}
    for doc in docs:
        iid = doc["item_id"]
        if iid not in result or doc.get("created_at", "") > result[iid].get("created_at", ""):
            result[iid] = serialize(doc)
    return result


async def bulk_set_ref_prices(
    quarter_id: str,
    prices: list[dict],  # [{item_id, unit, ref_price, notes}]
    *,
    user: dict,
) -> dict:
    """Bulk set reference prices for a quarter."""
    ok = 0
    errors = []
    for p in prices:
        try:
            await set_ref_price(
                quarter_id=quarter_id,
                item_id=p["item_id"],
                unit=p.get("unit", "pcs"),
                ref_price=float(p.get("ref_price", 0)),
                notes=p.get("notes"),
                user=user,
            )
            ok += 1
        except Exception as e:
            errors.append({"item_id": p.get("item_id"), "error": str(e)})
    return {"ok": ok, "errors": errors}


# =================== MARKET LIST VIEW ===================

async def get_market_list(
    *,
    quarter_id: Optional[str] = None,
    category_id: Optional[str] = None,
    search: Optional[str] = None,
    ml_status: Optional[str] = None,
    brand: Optional[str] = None,
    page: int = 1,
    per_page: int = 50,
) -> tuple[list[dict], dict]:
    """Get market list — items enriched with reference price for given quarter."""
    db = get_db()

    # Resolve quarter
    if not quarter_id:
        active = await get_active_quarter()
        quarter_id = active["id"] if active else None

    # Query items
    item_q: dict = {"deleted_at": None}
    if category_id:
        item_q["category_id"] = category_id
    if ml_status:
        item_q["ml_status"] = ml_status
    else:
        # Default: show active items + pending_review
        item_q["ml_status"] = {"$in": ["active", "pending_review", None]}
    if brand:
        item_q["brand_availability"] = {"$elemMatch": {"$eq": brand}}
    if search:
        item_q["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"code": {"$regex": search, "$options": "i"}},
        ]

    skip = (page - 1) * per_page
    items = await db.items.find(item_q).sort([("name", 1)]).skip(skip).limit(per_page).to_list(per_page)
    total = await db.items.count_documents(item_q)

    # Enrich with reference prices and category names
    item_ids = [item["id"] for item in items]
    category_ids = list({item.get("category_id") for item in items if item.get("category_id")})

    # Fetch ref prices for active quarter
    ref_price_map = await get_ref_prices_bulk(item_ids, quarter_id=quarter_id) if quarter_id else {}

    # Fetch previous quarter prices for comparison
    prev_ref_price_map: dict = {}
    if quarter_id:
        quarter_doc = await db.market_list_quarters.find_one({"id": quarter_id})
        if quarter_doc:
            prev_docs = await db.market_list_prices.find({
                "item_id": {"$in": item_ids},
                "quarter_id": {"$ne": quarter_id},
            }).sort([("created_at", -1)]).to_list(1000)
            seen = set()
            for doc in prev_docs:
                iid = doc["item_id"]
                if iid not in seen:
                    prev_ref_price_map[iid] = serialize(doc)
                    seen.add(iid)

    # Fetch categories
    cats = await db.categories.find({"id": {"$in": category_ids}}).to_list(100)
    cat_map = {c["id"]: c["name"] for c in cats}

    # Enrich
    enriched = []
    for item in items:
        d = serialize(item)
        ref = ref_price_map.get(item["id"])
        prev_ref = prev_ref_price_map.get(item["id"])
        d["ref_price"] = ref["ref_price"] if ref else None
        d["ref_price_unit"] = ref["unit"] if ref else None
        d["ref_quarter_label"] = ref["quarter_label"] if ref else None
        d["ref_variance_pct"] = ref.get("variance_pct") if ref else None
        d["prev_ref_price"] = prev_ref["ref_price"] if prev_ref else None
        d["prev_quarter_label"] = prev_ref["quarter_label"] if prev_ref else None
        d["category_name"] = cat_map.get(item.get("category_id", ""), "")
        d["ml_status"] = item.get("ml_status", "active")
        enriched.append(d)

    return enriched, {"page": page, "per_page": per_page, "total": total, "quarter_id": quarter_id}


async def approve_pending_item(
    item_id: str,
    category_id: str,
    ref_price: Optional[float] = None,
    *,
    user: dict,
) -> dict:
    """Approve a pending_review item: assign category, optionally set ref price."""
    db = get_db()
    item = await db.items.find_one({"id": item_id, "deleted_at": None})
    if not item:
        raise NotFoundError("Item tidak ditemukan")
    if item.get("ml_status") != "pending_review":
        raise ValidationError("Item bukan dalam status pending_review")
    # Validate category exists
    cat = await db.categories.find_one({"id": category_id, "deleted_at": None})
    if not cat:
        raise NotFoundError("Category tidak ditemukan")

    updates: dict = {
        "ml_status": "active",
        "category_id": category_id,
        "updated_at": _now(),
        "updated_by": user["id"],
    }
    await db.items.update_one({"id": item_id}, {"$set": updates})

    # Set ref price if provided
    if ref_price and ref_price > 0:
        active_q = await get_active_quarter()
        if active_q:
            unit = item.get("unit_default", "pcs")
            await set_ref_price(
                quarter_id=active_q["id"],
                item_id=item_id,
                unit=unit,
                ref_price=ref_price,
                notes=f"Set saat approval item pending oleh {user.get('email', user['id'])}",
                user=user,
            )

    fresh = await db.items.find_one({"id": item_id})
    logger.info(f"Approved pending item {item_id} by {user['id']}")
    return serialize(fresh)


async def get_all_quarter_prices_for_item(item_id: str) -> list[dict]:
    """Get ref prices for all quarters for a specific item."""
    db = get_db()
    docs = await db.market_list_prices.find({"item_id": item_id}).sort([
        ("quarter_id", 1)
    ]).to_list(50)
    result = [serialize(d) for d in docs]
    # Attach quarter label
    quarter_ids = [d["quarter_id"] for d in result]
    quarters = await db.market_list_quarters.find({"id": {"$in": quarter_ids}}).to_list(50)
    q_map = {q["id"]: q["label"] for q in quarters}
    for d in result:
        d["quarter_label"] = q_map.get(d["quarter_id"], d.get("quarter_label", ""))
    return sorted(result, key=lambda x: (x.get("quarter_label") or ""), reverse=True)


# =================== EXCEL EXPORT ===================

async def export_market_list_excel(year: int) -> bytes:
    """Export market list to Excel matching Torado Group format."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ValidationError("openpyxl tidak terinstall")

    db = get_db()

    # Get all quarters for this year (sorted Q1→Q4)
    quarters = await db.market_list_quarters.find({"year": year}).sort([("quarter", 1)]).to_list(10)
    quarter_ids = [q["id"] for q in quarters]
    quarter_labels = [q["label"] for q in quarters]

    # Get all active items
    items = await db.items.find({"deleted_at": None, "ml_status": {"$in": ["active", None]}}).sort([("name", 1)]).to_list(2000)
    item_ids = [item["id"] for item in items]

    # Get all prices for these quarters + items
    price_docs = await db.market_list_prices.find({
        "quarter_id": {"$in": quarter_ids},
        "item_id": {"$in": item_ids},
    }).to_list(20000)
    # Map: {(item_id, quarter_id): ref_price}
    price_map = {}
    for p in price_docs:
        price_map[(p["item_id"], p["quarter_id"])] = p.get("ref_price", 0)

    # Get categories
    cat_ids = list({item.get("category_id") for item in items if item.get("category_id")})
    cats = await db.categories.find({"id": {"$in": cat_ids}}).to_list(200)
    cat_map = {c["id"]: c["name"] for c in cats}

    # Get all brands for availability flags
    brands = await db.brands.find({"deleted_at": None}).sort([("name", 1)]).to_list(20)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"MASTER {year}"

    # Styles
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    price_fill = PatternFill("solid", fgColor="D6E4F0")
    pending_fill = PatternFill("solid", fgColor="FFF3CD")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # Build headers
    static_cols = ["NO", "ID", "REGIST DATE", "ITEMS", "UNIT (PROD)", "UNIT (COST)", "CONVERT UNIT", "CATEGORY"]
    price_cols = [f"PRICE ({ql})" for ql in quarter_labels]
    flag_cols = [b["code"].upper() for b in brands]
    all_cols = static_cols + price_cols + ["PREVIOUS PRICE", "VARIANCE (%)"] + flag_cols

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(all_cols))}1")
    title_cell = ws["A1"]
    title_cell.value = f"TORADO GROUP — MARKET LIST MASTER {year}"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = center_align

    # Subtitle
    ws.merge_cells(f"A2:{get_column_letter(len(all_cols))}2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated: {datetime.now().strftime('%d %B %Y')} | Aurora F&B ERP v0.3.0"
    sub_cell.font = Font(italic=True, size=9, color="666666")
    sub_cell.alignment = center_align

    # Header row (row 3)
    for col_idx, col_name in enumerate(all_cols, 1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill if col_name not in price_cols else PatternFill("solid", fgColor="2E75B6")
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows
    brand_codes = [b["code"].upper() for b in brands]
    for row_idx, item in enumerate(items, 4):
        item_brand_avail = [bc.upper() for bc in (item.get("brand_availability") or [])]
        conversion = ""
        cvs = item.get("conversion_units") or []
        if cvs and isinstance(cvs, list):
            parts = [f"1 {c.get('unit', '')} = {c.get('factor', '')} {item.get('unit_default', '')}" for c in cvs[:2]]
            conversion = "; ".join(parts)

        # Previous price = last price before the earliest quarter in this year
        prev_price = ""
        variance_pct = ""
        last_quarter_price = None
        for q_id in quarter_ids:
            p = price_map.get((item["id"], q_id))
            if p:
                last_quarter_price = p

        row_values = [
            row_idx - 3,                              # NO
            item.get("code", ""),                     # ID
            (item.get("created_at", "") or "")[:10],  # REGIST DATE
            item.get("name", ""),                      # ITEMS
            item.get("unit_default", ""),              # UNIT (PROD)
            item.get("unit_default", ""),              # UNIT (COST)
            conversion,                               # CONVERT UNIT
            cat_map.get(item.get("category_id", ""), ""),  # CATEGORY
        ]
        # Price columns per quarter
        for q_id in quarter_ids:
            p = price_map.get((item["id"], q_id))
            row_values.append(p if p else "")
        row_values.append(prev_price)  # PREVIOUS PRICE
        row_values.append(variance_pct)  # VARIANCE
        # Brand flags
        for bc in brand_codes:
            row_values.append(1 if bc in item_brand_avail else 0)

        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            col_name = all_cols[col_idx - 1]
            if col_name in price_cols:
                cell.fill = price_fill
                cell.number_format = '#,##0'
                cell.alignment = right_align
            elif col_name in ("PREVIOUS PRICE", "VARIANCE (%)"):
                cell.alignment = right_align
            elif col_name == "NO":
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            if item.get("ml_status") == "pending_review":
                cell.fill = pending_fill

    # Column widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 18
    for i in range(len(price_cols)):
        ws.column_dimensions[get_column_letter(9 + i)].width = 16
    # Freeze panes
    ws.freeze_panes = "E4"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
