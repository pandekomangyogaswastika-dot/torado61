"""Reports Excel — AP Aging Excel.

Auto-extracted from former monolithic reports_excel_finance_service.py.
"""
from __future__ import annotations

import logging
from typing import Optional



logger = logging.getLogger("aurora.reports")


async def generate_ap_aging_excel(
    *,
    as_of_date: Optional[str] = None,
    vendor_ids: Optional[list[str]] = None,
):
    """Generate AP Aging Excel report.
    
    Columns: Vendor, Current, 1-30 days, 31-60 days, 61-90 days, 90+ days, Total
    Features: Aging buckets, vendor detail breakdown, color coding for overdue
    """
    from services.excel_export_service import (
        create_workbook,
        add_report_header,
        write_table_headers,
        write_data_row,
        write_total_row,
        autosize_columns,
        apply_currency_format,
        format_date,
    )
    from openpyxl.styles import PatternFill, Font
    from services import finance_service
    
    # Use existing ap_aging service
    ap_data = await finance_service.ap_aging(as_of=as_of_date)
    
    rows = ap_data.get("rows", [])
    
    # Apply vendor filter
    if vendor_ids:
        rows = [r for r in rows if r.get("vendor_id") in vendor_ids]
    
    # Create workbook
    wb = create_workbook("AP Aging")
    ws = wb.active
    
    # Header
    as_of = ap_data.get("as_of", "")
    current_row = add_report_header(
        ws,
        title="Accounts Payable Aging Report",
        subtitle=f"As of: {format_date(as_of) if as_of else 'Current'}",
    )
    
    # === Section 1: Summary by Vendor ===
    summary_header = ws.cell(row=current_row, column=1)
    summary_header.value = "SUMMARY BY VENDOR"
    summary_header.font = Font(name="Manrope", size=12, bold=True, color="1C1510")
    summary_header.fill = PatternFill(start_color="F5E9D5", end_color="F5E9D5", fill_type="solid")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
    current_row += 1
    
    headers = ["Vendor", "Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "Total"]
    current_row = write_table_headers(ws, headers, current_row)
    data_start_row = current_row
    
    grand = {"current": 0.0, "d_30": 0.0, "d_60": 0.0, "d_90": 0.0, "d_90p": 0.0, "total": 0.0}
    
    for r in sorted(rows, key=lambda x: x.get("total", 0), reverse=True):
        row_data = [
            r.get("vendor_name", ""),
            float(r.get("current", 0) or 0),
            float(r.get("d_30", 0) or 0),
            float(r.get("d_60", 0) or 0),
            float(r.get("d_90", 0) or 0),
            float(r.get("d_90p", 0) or 0),
            float(r.get("total", 0) or 0),
        ]
        write_data_row(ws, row_data, current_row, number_cols=[2, 3, 4, 5, 6, 7])
        
        # Color code 90+ overdue (red), 61-90 (orange), 31-60 (yellow)
        if float(r.get("d_90p", 0) or 0) > 0:
            ws.cell(row=current_row, column=6).fill = PatternFill(
                start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        if float(r.get("d_90", 0) or 0) > 0:
            ws.cell(row=current_row, column=5).fill = PatternFill(
                start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        if float(r.get("d_60", 0) or 0) > 0:
            ws.cell(row=current_row, column=4).fill = PatternFill(
                start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        for k in grand.keys():
            grand[k] += float(r.get(k, 0) or 0)
        current_row += 1
    
    # Grand total row
    if rows:
        total_row = [
            "TOTAL",
            grand["current"], grand["d_30"], grand["d_60"], grand["d_90"], grand["d_90p"], grand["total"],
        ]
        write_total_row(ws, total_row, current_row, number_cols=[2, 3, 4, 5, 6, 7])
        current_row += 1
    
    # Styling for summary section
    for col in [2, 3, 4, 5, 6, 7]:
        apply_currency_format(ws, col, data_start_row, current_row - 1, currency="Rp")
    
    # === Section 2: Detail by Invoice ===
    current_row += 2  # Spacing
    
    detail_header = ws.cell(row=current_row, column=1)
    detail_header.value = "DETAIL BY INVOICE"
    detail_header.font = Font(name="Manrope", size=12, bold=True, color="1C1510")
    detail_header.fill = PatternFill(start_color="F5E9D5", end_color="F5E9D5", fill_type="solid")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    current_row += 1
    
    detail_headers = ["Vendor", "Doc No", "Invoice No", "Receive Date", "Due Date", "Days Overdue", "Bucket", "Amount"]
    current_row = write_table_headers(ws, detail_headers, current_row)
    detail_start_row = current_row
    
    detail_total = 0.0
    for r in rows:
        vendor_name = r.get("vendor_name", "")
        for it in r.get("items", []):
            bucket_label = {
                "current": "Current",
                "d_30": "1-30 Days",
                "d_60": "31-60 Days",
                "d_90": "61-90 Days",
                "d_90p": "90+ Days",
            }.get(it.get("bucket"), it.get("bucket", ""))
            
            row_data = [
                vendor_name,
                it.get("doc_no", ""),
                it.get("invoice_no", ""),
                format_date(it.get("receive_date")),
                format_date(it.get("due_date")),
                int(it.get("days_overdue", 0) or 0),
                bucket_label,
                float(it.get("amount", 0) or 0),
            ]
            write_data_row(ws, row_data, current_row, number_cols=[6, 8])
            
            # Color code by bucket
            if it.get("bucket") == "d_90p":
                ws.cell(row=current_row, column=7).fill = PatternFill(
                    start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            elif it.get("bucket") == "d_90":
                ws.cell(row=current_row, column=7).fill = PatternFill(
                    start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
            elif it.get("bucket") == "d_60":
                ws.cell(row=current_row, column=7).fill = PatternFill(
                    start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            detail_total += float(it.get("amount", 0) or 0)
            current_row += 1
    
    # Detail total row
    if detail_total > 0:
        total_row = ["", "", "", "", "", "", "TOTAL", detail_total]
        write_total_row(ws, total_row, current_row, number_cols=[8])
        current_row += 1
    
    # Detail styling
    apply_currency_format(ws, 8, detail_start_row, current_row - 1, currency="Rp")
    autosize_columns(ws)
    
    return wb


# ============================================================
# 11. PHASE 4.5 — REPORT BUILDER EXCEL EXPORT
# ============================================================
