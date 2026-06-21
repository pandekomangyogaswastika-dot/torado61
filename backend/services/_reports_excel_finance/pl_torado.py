"""Reports Excel — P&L Torado Excel.

Auto-extracted from former monolithic reports_excel_finance_service.py.
"""
from __future__ import annotations

import logging
from typing import Optional

from core.db import get_db
from core.exceptions import ValidationError


logger = logging.getLogger("aurora.reports")


async def generate_pl_torado_excel(
    *,
    period_from: str,  # YYYY-MM
    period_to: str,    # YYYY-MM
    dim_outlet: Optional[str] = None,
):
    """Generate Profit & Loss in Torado custom format.
    
    Layout (matches legacy Excel):
      Rows: Account hierarchy (Revenue, COGS, Gross Profit, OPEX, Net Income)
      Cols: Period months (Jan, Feb, ..., Dec) + YTD total
    
    Sections:
      A. REVENUE (cr-normal)
        - per revenue COA row × each month col
        - Subtotal Revenue
      B. COGS (dr-normal)
        - per cogs COA row × each month col
        - Subtotal COGS
      C. GROSS PROFIT = Revenue - COGS  (row, with margin %)
      D. OPERATING EXPENSE (dr-normal expense)
        - per expense COA row × each month col
        - Subtotal Expense
      E. NET INCOME = Gross Profit - Expense (row, with margin %)
    """
    from services.excel_export_service import (
        create_workbook,
        add_report_header,
        freeze_header_row,
        TORADO_ESPRESSO,
        TORADO_GOLD,
        TORADO_LIGHT_TEXT,
        BORDER_THIN,
    )
    from openpyxl.styles import Alignment, Font, PatternFill
    from services import finance_service
    
    # Validate periods
    try:
        y1, m1 = [int(x) for x in period_from.split("-")]
        y2, m2 = [int(x) for x in period_to.split("-")]
        assert 1 <= m1 <= 12 and 1 <= m2 <= 12
    except Exception as e:
        raise ValidationError("period_from / period_to harus YYYY-MM") from e
    
    if (y1, m1) > (y2, m2):
        raise ValidationError("period_from harus <= period_to")
    
    # Build list of months
    months: list[str] = []
    cy, cm = y1, m1
    while (cy, cm) <= (y2, m2):
        months.append(f"{cy:04d}-{cm:02d}")
        cm += 1
        if cm > 12:
            cm = 1
            cy += 1
        if len(months) > 24:
            raise ValidationError("Maksimal 24 bulan per export")
    
    # Run P&L for each month
    pl_per_month: dict[str, dict] = {}
    for p in months:
        pl_per_month[p] = await finance_service.profit_loss(
            period=p, dim_outlet=dim_outlet, compare_prev=False,
        )
    
    # Outlet name for header
    db = get_db()
    outlet_name = ""
    if dim_outlet:
        o = await db.outlets.find_one({"id": dim_outlet, "deleted_at": None})
        if o:
            outlet_name = o.get("name", dim_outlet)
    
    # Collect all unique COAs that appear in any month, grouped by section
    coas_by_section: dict[str, dict] = {"revenue": {}, "cogs": {}, "expense": {}}
    for p, pl in pl_per_month.items():
        for sec_key, sec_rows in pl.get("sections", {}).items():
            for r in sec_rows:
                cid = r["coa_id"]
                if cid not in coas_by_section[sec_key]:
                    coas_by_section[sec_key][cid] = {
                        "coa_id": cid, "code": r["code"], "name": r["name"],
                        "amounts": {},
                    }
                coas_by_section[sec_key][cid]["amounts"][p] = float(r.get("amount", 0) or 0)
    
    # Sort each section by COA code
    for sec_key in coas_by_section:
        coas_by_section[sec_key] = sorted(
            coas_by_section[sec_key].values(),
            key=lambda r: r.get("code") or "",
        )
    
    # Workbook
    wb = create_workbook("P&L Torado")
    ws = wb.active
    
    subtitle_parts = [f"Period: {period_from} — {period_to}"]
    if outlet_name:
        subtitle_parts.append(f"Outlet: {outlet_name}")
    else:
        subtitle_parts.append("Outlet: Konsolidasi")
    
    current_row = add_report_header(
        ws,
        title="Profit & Loss Statement (Torado Format)",
        subtitle=" | ".join(subtitle_parts),
    )
    
    # === Header row: Account | Code | Month1 | Month2 | ... | YTD ===
    n_months = len(months)
    total_cols = 2 + n_months + 1  # Account + Code + months + YTD
    
    # Header styling
    HDR_FILL = PatternFill(start_color=TORADO_ESPRESSO, end_color=TORADO_ESPRESSO, fill_type="solid")
    HDR_FONT = Font(name="Manrope", size=11, bold=True, color=TORADO_LIGHT_TEXT)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    headers = ["Account", "Code"] + months + ["YTD"]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=idx)
        cell.value = h
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = BORDER_THIN
    current_row += 1
    data_start_row = current_row
    
    # Number cells styling helpers
    NUM_ALIGN = Alignment(horizontal="right", vertical="center")
    LBL_ALIGN = Alignment(horizontal="left", vertical="center")
    
    def write_label(row, col, value, *, bold=False, fill_color=None, text_color="000000", indent=0):
        cell = ws.cell(row=row, column=col)
        prefix = "  " * indent
        cell.value = f"{prefix}{value}"
        cell.font = Font(name="Manrope", size=10, bold=bold, color=text_color)
        cell.alignment = LBL_ALIGN
        cell.border = BORDER_THIN
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    def write_number(row, col, value, *, bold=False, fill_color=None, text_color="000000"):
        cell = ws.cell(row=row, column=col)
        cell.value = float(value or 0)
        cell.font = Font(name="IBM Plex Mono", size=10, bold=bold, color=text_color)
        cell.alignment = NUM_ALIGN
        cell.number_format = '"Rp"#,##0;[Red]("Rp"#,##0)'
        cell.border = BORDER_THIN
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    def write_section_header(row, label, fill_color=TORADO_GOLD, text_color=TORADO_ESPRESSO):
        for col in range(1, total_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.value = label if col == 1 else None
            cell.font = Font(name="Manrope", size=11, bold=True, color=text_color)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.alignment = LBL_ALIGN
            cell.border = BORDER_THIN
    
    # === SECTION A: REVENUE ===
    write_section_header(current_row, "A. REVENUE")
    current_row += 1
    
    rev_subtotal_per_month: dict[str, float] = {p: 0.0 for p in months}
    for r in coas_by_section["revenue"]:
        write_label(current_row, 1, r["name"], indent=1)
        write_label(current_row, 2, r["code"])
        for idx, p in enumerate(months):
            v = r["amounts"].get(p, 0.0)
            write_number(current_row, 3 + idx, v)
            rev_subtotal_per_month[p] += v
        write_number(current_row, 3 + n_months, sum(r["amounts"].values()))
        current_row += 1
    
    # Revenue subtotal
    write_label(current_row, 1, "Total Revenue", bold=True, fill_color="F0E5C8")
    write_label(current_row, 2, "", bold=True, fill_color="F0E5C8")
    for idx, p in enumerate(months):
        write_number(current_row, 3 + idx, rev_subtotal_per_month[p], bold=True, fill_color="F0E5C8")
    write_number(current_row, 3 + n_months, sum(rev_subtotal_per_month.values()),
                 bold=True, fill_color="F0E5C8")
    current_row += 1
    current_row += 1  # spacing
    
    # === SECTION B: COGS ===
    write_section_header(current_row, "B. COST OF GOODS SOLD")
    current_row += 1
    
    cogs_subtotal_per_month: dict[str, float] = {p: 0.0 for p in months}
    for r in coas_by_section["cogs"]:
        write_label(current_row, 1, r["name"], indent=1)
        write_label(current_row, 2, r["code"])
        for idx, p in enumerate(months):
            v = r["amounts"].get(p, 0.0)
            write_number(current_row, 3 + idx, v)
            cogs_subtotal_per_month[p] += v
        write_number(current_row, 3 + n_months, sum(r["amounts"].values()))
        current_row += 1
    
    write_label(current_row, 1, "Total COGS", bold=True, fill_color="F0E5C8")
    write_label(current_row, 2, "", bold=True, fill_color="F0E5C8")
    for idx, p in enumerate(months):
        write_number(current_row, 3 + idx, cogs_subtotal_per_month[p], bold=True, fill_color="F0E5C8")
    write_number(current_row, 3 + n_months, sum(cogs_subtotal_per_month.values()),
                 bold=True, fill_color="F0E5C8")
    current_row += 1
    current_row += 1  # spacing
    
    # === SECTION C: GROSS PROFIT (Revenue - COGS) ===
    write_section_header(current_row, "C. GROSS PROFIT (A − B)", fill_color="C9E5C5", text_color="1A4D1F")
    current_row += 1
    
    gp_per_month: dict[str, float] = {}
    for p in months:
        gp_per_month[p] = rev_subtotal_per_month[p] - cogs_subtotal_per_month[p]
    
    write_label(current_row, 1, "Gross Profit", bold=True, fill_color="DCEFD8")
    write_label(current_row, 2, "", bold=True, fill_color="DCEFD8")
    for idx, p in enumerate(months):
        write_number(current_row, 3 + idx, gp_per_month[p], bold=True, fill_color="DCEFD8")
    write_number(current_row, 3 + n_months, sum(gp_per_month.values()),
                 bold=True, fill_color="DCEFD8")
    current_row += 1
    
    # Gross Margin %
    write_label(current_row, 1, "Gross Margin %", indent=1, fill_color="DCEFD8")
    write_label(current_row, 2, "", fill_color="DCEFD8")
    for idx, p in enumerate(months):
        rev = rev_subtotal_per_month[p]
        gp = gp_per_month[p]
        margin = (gp / rev * 100) if rev else 0
        cell = ws.cell(row=current_row, column=3 + idx)
        cell.value = margin / 100
        cell.font = Font(name="IBM Plex Mono", size=10, italic=True, color="555555")
        cell.alignment = NUM_ALIGN
        cell.number_format = "0.00%"
        cell.fill = PatternFill(start_color="DCEFD8", end_color="DCEFD8", fill_type="solid")
        cell.border = BORDER_THIN
    # YTD margin
    total_rev = sum(rev_subtotal_per_month.values())
    total_gp = sum(gp_per_month.values())
    cell = ws.cell(row=current_row, column=3 + n_months)
    cell.value = (total_gp / total_rev) if total_rev else 0
    cell.font = Font(name="IBM Plex Mono", size=10, italic=True, color="555555")
    cell.alignment = NUM_ALIGN
    cell.number_format = "0.00%"
    cell.fill = PatternFill(start_color="DCEFD8", end_color="DCEFD8", fill_type="solid")
    cell.border = BORDER_THIN
    current_row += 1
    current_row += 1
    
    # === SECTION D: OPERATING EXPENSE ===
    write_section_header(current_row, "D. OPERATING EXPENSE")
    current_row += 1
    
    exp_subtotal_per_month: dict[str, float] = {p: 0.0 for p in months}
    for r in coas_by_section["expense"]:
        write_label(current_row, 1, r["name"], indent=1)
        write_label(current_row, 2, r["code"])
        for idx, p in enumerate(months):
            v = r["amounts"].get(p, 0.0)
            write_number(current_row, 3 + idx, v)
            exp_subtotal_per_month[p] += v
        write_number(current_row, 3 + n_months, sum(r["amounts"].values()))
        current_row += 1
    
    write_label(current_row, 1, "Total OPEX", bold=True, fill_color="F0E5C8")
    write_label(current_row, 2, "", bold=True, fill_color="F0E5C8")
    for idx, p in enumerate(months):
        write_number(current_row, 3 + idx, exp_subtotal_per_month[p], bold=True, fill_color="F0E5C8")
    write_number(current_row, 3 + n_months, sum(exp_subtotal_per_month.values()),
                 bold=True, fill_color="F0E5C8")
    current_row += 1
    current_row += 1
    
    # === SECTION E: NET INCOME (GP - OPEX) ===
    write_section_header(current_row, "E. NET INCOME (C − D)", fill_color="C9E5C5", text_color="1A4D1F")
    current_row += 1
    
    ni_per_month: dict[str, float] = {}
    for p in months:
        ni_per_month[p] = gp_per_month[p] - exp_subtotal_per_month[p]
    
    write_label(current_row, 1, "Net Income", bold=True, fill_color="DCEFD8")
    write_label(current_row, 2, "", bold=True, fill_color="DCEFD8")
    for idx, p in enumerate(months):
        ni = ni_per_month[p]
        text_color = "1A7F37" if ni >= 0 else "B91C1C"
        write_number(current_row, 3 + idx, ni, bold=True, fill_color="DCEFD8", text_color=text_color)
    total_ni = sum(ni_per_month.values())
    write_number(current_row, 3 + n_months, total_ni, bold=True,
                 fill_color="DCEFD8",
                 text_color="1A7F37" if total_ni >= 0 else "B91C1C")
    current_row += 1
    
    # Net Margin %
    write_label(current_row, 1, "Net Margin %", indent=1, fill_color="DCEFD8")
    write_label(current_row, 2, "", fill_color="DCEFD8")
    for idx, p in enumerate(months):
        rev = rev_subtotal_per_month[p]
        ni = ni_per_month[p]
        margin = (ni / rev) if rev else 0
        cell = ws.cell(row=current_row, column=3 + idx)
        cell.value = margin
        cell.font = Font(name="IBM Plex Mono", size=10, italic=True, color="555555")
        cell.alignment = NUM_ALIGN
        cell.number_format = "0.00%"
        cell.fill = PatternFill(start_color="DCEFD8", end_color="DCEFD8", fill_type="solid")
        cell.border = BORDER_THIN
    cell = ws.cell(row=current_row, column=3 + n_months)
    cell.value = (total_ni / total_rev) if total_rev else 0
    cell.font = Font(name="IBM Plex Mono", size=10, italic=True, color="555555")
    cell.alignment = NUM_ALIGN
    cell.number_format = "0.00%"
    cell.fill = PatternFill(start_color="DCEFD8", end_color="DCEFD8", fill_type="solid")
    cell.border = BORDER_THIN
    current_row += 1
    
    # Set column widths (Account wide, Code narrow, months medium, YTD wider)
    from openpyxl.utils import get_column_letter
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 10
    for i in range(n_months):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14
    ws.column_dimensions[get_column_letter(3 + n_months)].width = 16
    
    freeze_header_row(ws, header_row=data_start_row - 1)
    
    return wb
