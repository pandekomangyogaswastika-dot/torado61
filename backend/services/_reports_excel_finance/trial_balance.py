"""Reports Excel — Trial Balance Excel.

Auto-extracted from former monolithic reports_excel_finance_service.py.
"""
from __future__ import annotations

import logging
from typing import Optional

from core.db import get_db


logger = logging.getLogger("aurora.reports")


async def generate_trial_balance_excel(
    *,
    period: str,
    dim_outlet: Optional[str] = None,
):
    """Generate Trial Balance Excel report for a given period (YYYY-MM).
    
    Columns: COA Code, COA Name, Type, Normal Balance, Opening, Period Dr, Period Cr, Closing
    Features: Balance check (Dr=Cr), grouping by type
    """
    from services.excel_export_service import (
        create_workbook,
        add_report_header,
        write_table_headers,
        write_data_row,
        write_total_row,
        autosize_columns,
        freeze_header_row,
        apply_currency_format,
    )
    from openpyxl.styles import PatternFill, Font, Alignment
    from services import finance_service
    
    # Use existing trial_balance service
    tb_data = await finance_service.trial_balance(period=period, dim_outlet=dim_outlet)
    
    db = get_db()
    
    # Get outlet name if filtered
    outlet_name = ""
    if dim_outlet:
        outlet_doc = await db.outlets.find_one({"id": dim_outlet, "deleted_at": None})
        if outlet_doc:
            outlet_name = outlet_doc.get("name", dim_outlet)
    
    # Create workbook
    wb = create_workbook("Trial Balance")
    ws = wb.active
    
    # Header
    subtitle_parts = [f"Period: {period}"]
    if outlet_name:
        subtitle_parts.append(f"Outlet: {outlet_name}")
    
    totals = tb_data.get("totals", {})
    is_balanced = totals.get("is_balanced_period", False)
    balance_status = "BALANCED ✓" if is_balanced else "UNBALANCED ✗"
    subtitle_parts.append(f"Status: {balance_status}")
    
    current_row = add_report_header(
        ws,
        title="Trial Balance Report",
        subtitle=" | ".join(subtitle_parts),
    )
    
    # Headers
    headers = ["COA Code", "COA Name", "Type", "Normal", "Opening", "Period Dr", "Period Cr", "Closing"]
    current_row = write_table_headers(ws, headers, current_row)
    data_start_row = current_row
    
    rows = tb_data.get("rows", [])
    
    # Group rows by type for organized display
    type_order = ["asset", "liability", "equity", "revenue", "cogs", "expense", "other"]
    grouped: dict[str, list] = {}
    for r in rows:
        t = (r.get("type") or "other").lower()
        grouped.setdefault(t, []).append(r)
    
    for t in type_order:
        type_rows = grouped.get(t, [])
        if not type_rows:
            continue
        
        # Section header row
        section_cell = ws.cell(row=current_row, column=1)
        section_cell.value = t.upper()
        section_cell.font = Font(name="Manrope", size=11, bold=True, color="1C1510")
        section_cell.fill = PatternFill(start_color="F5E9D5", end_color="F5E9D5", fill_type="solid")
        section_cell.alignment = Alignment(horizontal="left", vertical="center")
        # Merge across columns for visual grouping
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
        current_row += 1
        
        # Data rows
        for r in sorted(type_rows, key=lambda x: x.get("code") or ""):
            row_data = [
                r.get("code", ""),
                r.get("name", ""),
                r.get("type", ""),
                r.get("normal_balance", ""),
                float(r.get("opening", 0)),
                float(r.get("period_dr", 0)),
                float(r.get("period_cr", 0)),
                float(r.get("closing", 0)),
            ]
            write_data_row(ws, row_data, current_row, number_cols=[5, 6, 7, 8])
            current_row += 1
    
    # Total row
    total_row = [
        "", "", "", "TOTAL",
        float(totals.get("opening_dr", 0)) - float(totals.get("opening_cr", 0)),
        float(totals.get("period_dr", 0)),
        float(totals.get("period_cr", 0)),
        float(totals.get("closing_dr", 0)) - float(totals.get("closing_cr", 0)),
    ]
    write_total_row(ws, total_row, current_row, number_cols=[5, 6, 7, 8])
    current_row += 1
    
    # Balance check row
    balance_cell = ws.cell(row=current_row, column=1)
    balance_cell.value = f"Status: Period Dr (Rp {totals.get('period_dr', 0):,.0f}) {'=' if is_balanced else '!='} Period Cr (Rp {totals.get('period_cr', 0):,.0f})"
    balance_cell.font = Font(name="Manrope", size=10, bold=True, italic=True,
                              color="1A7F37" if is_balanced else "B91C1C")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    
    # Styling
    apply_currency_format(ws, 5, data_start_row, current_row - 1, currency="Rp")
    apply_currency_format(ws, 6, data_start_row, current_row - 1, currency="Rp")
    apply_currency_format(ws, 7, data_start_row, current_row - 1, currency="Rp")
    apply_currency_format(ws, 8, data_start_row, current_row - 1, currency="Rp")
    autosize_columns(ws)
    freeze_header_row(ws, header_row=data_start_row - 1)
    
    return wb
