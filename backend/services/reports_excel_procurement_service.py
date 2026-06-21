"""reports_excel_procurement_service — Excel export functions."""
from __future__ import annotations
import logging
from datetime import datetime, timezone
from typing import Optional
from core.db import get_db
from core.exceptions import ValidationError
logger = logging.getLogger("aurora.reports")

def _now():
    return datetime.now(timezone.utc).isoformat()

def _parse_date(s, *, fallback_today=False):
    if not s:
        return datetime.now(timezone.utc) if fallback_today else None
    try:
        return datetime.strptime(s, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except ValueError as e:
        raise ValidationError(f"Invalid date format: {s}") from e

async def generate_po_summary_excel(
    *,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    vendor_ids: Optional[list[str]] = None,
    status: Optional[str] = None,
):
    """Generate PO Summary Excel report.
    
    Columns: PO No, Date, Vendor, Status, Total Amount, Items Count, Delivery Date
    Features: Status color coding, vendor grouping optional
    """
    from services.excel_export_service import (
        create_workbook,
        add_report_header,
        write_table_headers,
        write_data_row,
        write_total_row,
        autosize_columns,
        freeze_header_row,
        apply_currency_format,
        format_date,
    )
    from openpyxl.styles import PatternFill
    
    db = get_db()
    
    # Build query
    match_filter: dict = {"deleted_at": None}
    if date_from or date_to:
        match_filter["po_date"] = {}
        if date_from:
            match_filter["po_date"]["$gte"] = date_from
        if date_to:
            match_filter["po_date"]["$lte"] = date_to
    if vendor_ids:
        match_filter["vendor_id"] = {"$in": vendor_ids}
    if status:
        match_filter["status"] = status
    
    # Load POs
    pos = []
    async for doc in db.purchase_orders.find(match_filter).sort([("po_date", -1), ("created_at", -1)]).limit(500):
        pos.append(doc)
    
    # Load vendors
    vendor_map = {}
    async for v in db.vendors.find({"deleted_at": None}):
        vendor_map[v["id"]] = v.get("name", v["id"])
    
    # Create workbook
    wb = create_workbook("PO Summary")
    ws = wb.active
    
    # Add header
    subtitle = f"Period: {format_date(date_from) if date_from else 'All'} - {format_date(date_to) if date_to else 'All'}"
    current_row = add_report_header(
        ws,
        title="Purchase Order Summary Report",
        subtitle=subtitle,
    )
    
    # Table headers
    headers = ["PO No", "PO Date", "Vendor", "Status", "Items Count", "Total Amount", "Delivery Date"]
    current_row = write_table_headers(ws, headers, current_row)
    data_start_row = current_row
    
    # Data rows
    grand_total = 0.0
    
    for po in pos:
        vendor_name = vendor_map.get(po.get("vendor_id"), po.get("vendor_id") or "")
        items_count = len(po.get("lines", []))
        total_amount = float(po.get("total_amount", 0) or 0)
        status_val = po.get("status", "draft")
        
        row_data = [
            po.get("po_no", ""),
            format_date(po.get("po_date")),
            vendor_name,
            status_val,
            items_count,
            total_amount,
            format_date(po.get("delivery_date")),
        ]
        write_data_row(ws, row_data, current_row, number_cols=[5, 6])
        
        # Color code status cell
        status_cell = ws.cell(row=current_row, column=4)
        if status_val == "approved":
            status_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        elif status_val == "rejected":
            status_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        elif status_val == "pending":
            status_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        
        grand_total += total_amount
        current_row += 1
    
    # Total row
    if pos:
        total_row = ["", "", "", "TOTAL", "", grand_total, ""]
        write_total_row(ws, total_row, current_row, number_cols=[6])
    
    # Styling
    apply_currency_format(ws, 6, data_start_row, current_row, currency="Rp")
    autosize_columns(ws)
    freeze_header_row(ws, header_row=data_start_row - 1)
    
    return wb


async def generate_gr_summary_excel(
    *,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    vendor_ids: Optional[list[str]] = None,
    po_ids: Optional[list[str]] = None,
):
    """Generate GR Summary Excel report.
    
    Columns: GR No, Date, PO No, Vendor, Status, Items Count, Total Amount, Variance
    Features: Variance analysis (PO vs GR amounts)
    """
    from services.excel_export_service import (
        create_workbook,
        add_report_header,
        write_table_headers,
        write_data_row,
        write_total_row,
        autosize_columns,
        freeze_header_row,
        apply_currency_format,
        format_date,
    )
    
    db = get_db()
    
    # Build query
    match_filter: dict = {"deleted_at": None}
    if date_from or date_to:
        match_filter["gr_date"] = {}
        if date_from:
            match_filter["gr_date"]["$gte"] = date_from
        if date_to:
            match_filter["gr_date"]["$lte"] = date_to
    if vendor_ids:
        match_filter["vendor_id"] = {"$in": vendor_ids}
    if po_ids:
        match_filter["po_id"] = {"$in": po_ids}
    
    # Load GRs
    grs = []
    async for doc in db.goods_receipts.find(match_filter).sort([("gr_date", -1), ("created_at", -1)]).limit(500):
        grs.append(doc)
    
    # Load vendors & POs
    vendor_map = {}
    async for v in db.vendors.find({"deleted_at": None}):
        vendor_map[v["id"]] = v.get("name", v["id"])
    
    po_map = {}
    async for po in db.purchase_orders.find({"deleted_at": None}):
        po_map[po["id"]] = {
            "po_no": po.get("po_no", ""),
            "total_amount": float(po.get("total_amount", 0) or 0),
        }
    
    # Create workbook
    wb = create_workbook("GR Summary")
    ws = wb.active
    
    # Add header
    subtitle = f"Period: {format_date(date_from) if date_from else 'All'} - {format_date(date_to) if date_to else 'All'}"
    current_row = add_report_header(
        ws,
        title="Goods Receipt Summary Report",
        subtitle=subtitle,
    )
    
    # Table headers
    headers = ["GR No", "GR Date", "PO No", "Vendor", "Items Count", "GR Amount", "PO Amount", "Variance"]
    current_row = write_table_headers(ws, headers, current_row)
    data_start_row = current_row
    
    # Data rows
    grand_gr_total = 0.0
    grand_po_total = 0.0
    
    for gr in grs:
        vendor_name = vendor_map.get(gr.get("vendor_id"), gr.get("vendor_id") or "")
        po_info = po_map.get(gr.get("po_id"), {})
        po_no = po_info.get("po_no", "")
        po_amount = po_info.get("total_amount", 0)
        
        items_count = len(gr.get("lines", []))
        gr_amount = float(gr.get("total_amount", 0) or 0)
        variance = gr_amount - po_amount
        
        row_data = [
            gr.get("gr_no", ""),
            format_date(gr.get("gr_date")),
            po_no,
            vendor_name,
            items_count,
            gr_amount,
            po_amount,
            variance,
        ]
        write_data_row(ws, row_data, current_row, number_cols=[5, 6, 7, 8])
        
        grand_gr_total += gr_amount
        grand_po_total += po_amount
        current_row += 1
    
    # Total row
    if grs:
        total_variance = grand_gr_total - grand_po_total
        total_row = ["", "", "", "TOTAL", "", grand_gr_total, grand_po_total, total_variance]
        write_total_row(ws, total_row, current_row, number_cols=[6, 7, 8])
    
    # Styling
    apply_currency_format(ws, 6, data_start_row, current_row, currency="Rp")
    apply_currency_format(ws, 7, data_start_row, current_row, currency="Rp")
    apply_currency_format(ws, 8, data_start_row, current_row, currency="Rp")
    autosize_columns(ws)
    freeze_header_row(ws, header_row=data_start_row - 1)
    
    return wb


async def generate_vendor_performance_excel(
    *,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    vendor_ids: Optional[list[str]] = None,
):
    """Generate Vendor Performance Excel report.
    
    Leverage existing vendor_scorecard data + add Excel formatting.
    Columns: Vendor, PO Count, GR Count, Total Purchase Value, On-Time Delivery %, Avg Price Stability
    Features: Performance metrics + bar chart
    """
    from services.excel_export_service import (
        create_workbook,
        add_report_header,
        write_table_headers,
        write_data_row,
        write_total_row,
        autosize_columns,
        freeze_header_row,
        apply_currency_format,
        format_date,
        add_bar_chart,
    )
    
    db = get_db()
    
    # Build aggregation for vendor performance
    match_filter: dict = {"deleted_at": None}
    if date_from or date_to:
        match_filter["po_date"] = {}
        if date_from:
            match_filter["po_date"]["$gte"] = date_from
        if date_to:
            match_filter["po_date"]["$lte"] = date_to
    if vendor_ids:
        match_filter["vendor_id"] = {"$in": vendor_ids}
    
    # Aggregate PO stats per vendor
    pipeline = [
        {"$match": match_filter},
        {
            "$group": {
                "_id": "$vendor_id",
                "po_count": {"$sum": 1},
                "total_value": {"$sum": "$total_amount"},
            }
        },
        {"$sort": {"total_value": -1}},
    ]
    
    vendor_stats = []
    async for doc in db.purchase_orders.aggregate(pipeline):
        vendor_stats.append(doc)
    
    # Aggregate GR stats per vendor
    gr_filter: dict = {"deleted_at": None}
    if date_from or date_to:
        gr_filter["gr_date"] = {}
        if date_from:
            gr_filter["gr_date"]["$gte"] = date_from
        if date_to:
            gr_filter["gr_date"]["$lte"] = date_to
    if vendor_ids:
        gr_filter["vendor_id"] = {"$in": vendor_ids}
    
    gr_pipeline = [
        {"$match": gr_filter},
        {
            "$group": {
                "_id": "$vendor_id",
                "gr_count": {"$sum": 1},
            }
        },
    ]
    
    gr_stats = {}
    async for doc in db.goods_receipts.aggregate(gr_pipeline):
        gr_stats[doc["_id"]] = doc.get("gr_count", 0)
    
    # Load vendor names
    vendor_map = {}
    async for v in db.vendors.find({"deleted_at": None}):
        vendor_map[v["id"]] = v.get("name", v["id"])
    
    # Create workbook
    wb = create_workbook("Vendor Performance")
    ws = wb.active
    
    # Add header
    subtitle = f"Period: {format_date(date_from) if date_from else 'All'} - {format_date(date_to) if date_to else 'All'}"
    current_row = add_report_header(
        ws,
        title="Vendor Performance Report",
        subtitle=subtitle,
    )
    
    # Table headers
    headers = ["Vendor", "PO Count", "GR Count", "Total Purchase Value", "Fulfillment Rate"]
    current_row = write_table_headers(ws, headers, current_row)
    data_start_row = current_row
    
    # Data rows
    grand_po_count = 0
    grand_gr_count = 0
    grand_value = 0.0
    
    for stat in vendor_stats:
        vendor_id = stat["_id"]
        vendor_name = vendor_map.get(vendor_id, vendor_id or "")
        po_count = int(stat.get("po_count", 0) or 0)
        gr_count = gr_stats.get(vendor_id, 0)
        total_value = float(stat.get("total_value", 0) or 0)
        fulfillment_rate = (gr_count / po_count * 100) if po_count > 0 else 0
        
        row_data = [
            vendor_name,
            po_count,
            gr_count,
            total_value,
            f"{fulfillment_rate:.1f}%",
        ]
        write_data_row(ws, row_data, current_row, number_cols=[2, 3, 4])
        
        grand_po_count += po_count
        grand_gr_count += gr_count
        grand_value += total_value
        current_row += 1
    
    # Total row
    if vendor_stats:
        overall_fulfillment = (grand_gr_count / grand_po_count * 100) if grand_po_count > 0 else 0
        total_row = ["TOTAL", grand_po_count, grand_gr_count, grand_value, f"{overall_fulfillment:.1f}%"]
        write_total_row(ws, total_row, current_row, number_cols=[2, 3, 4])
    
    # Styling
    apply_currency_format(ws, 4, data_start_row, current_row, currency="Rp")
    autosize_columns(ws)
    freeze_header_row(ws, header_row=data_start_row - 1)
    
    # Add bar chart (Total Purchase Value)
    if vendor_stats and len(vendor_stats) > 1:
        data_range = f"D{data_start_row}:D{current_row - 1}"
        cats_range = f"A{data_start_row}:A{current_row - 1}"
        add_bar_chart(ws, "Total Purchase Value by Vendor", data_range, cats_range, position="G2")
    
    return wb


# ============================================================
# 10. FINANCE REPORTS (Phase 4.4)
# ============================================================

