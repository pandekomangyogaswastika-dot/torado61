"""Budget CSV/Excel import and Excel template generation."""
from __future__ import annotations

import csv
import io

from core.db import get_db
from models.budget import make_budget_doc, BUDGET_CATEGORIES

from services._budget._common import guess_category, MONTH_COLS, MONTH_IDX


async def import_csv(
    csv_content: str,
    period: str,
    *,
    outlet_id: str | None = None,
    scope: str = "outlet",
    user_id: str,
) -> dict:
    """Import budget from CSV. Columns: coa_code, amount, category (optional)."""
    db = get_db()
    reader = csv.DictReader(io.StringIO(csv_content))
    lines = []
    errors = []
    for i, row in enumerate(reader, 1):
        coa_code = (row.get("coa_code") or row.get("code") or "").strip()
        amount_str = (row.get("amount") or "0").strip().replace(",", "")
        try:
            amount = float(amount_str)
        except ValueError:
            errors.append(f"Row {i}: invalid amount '{amount_str}'")
            continue
        coa = await db.chart_of_accounts.find_one({"code": coa_code, "deleted_at": None})
        if not coa:
            errors.append(f"Row {i}: COA code '{coa_code}' not found")
            continue
        lines.append({
            "coa_id": coa["id"],
            "coa_code": coa_code,
            "coa_name": coa["name"],
            "category": row.get("category") or guess_category(coa_code),
            "amount": amount,
        })
    if errors and not lines:
        return {"success": False, "errors": errors, "imported": 0}
    doc = make_budget_doc(
        name=f"Budget Import {period}",
        period=period,
        period_type="monthly",
        scope=scope,
        outlet_id=outlet_id,
        brand_id=None,
        lines=lines,
        notes="Imported from CSV",
        created_by=user_id,
    )
    await db.budgets.insert_one(doc)
    return {"success": True, "imported": len(lines), "errors": errors, "budget_id": doc["id"]}


async def import_excel(
    file_bytes: bytes,
    period: str,
    *,
    outlet_id: str | None = None,
    brand_id: str | None = None,
    scope: str = "outlet",
    user_id: str,
) -> dict:
    """Import budget from Excel (.xlsx).

    Supports TWO templates:
    A. Simple:  coa_code | coa_name | amount | category
    B. Monthly: coa_code | coa_name | jan | feb | ... | dec | total | category
    """
    import io as _io
    try:
        import openpyxl
    except ImportError:
        return {"success": False, "errors": ["openpyxl not installed"], "imported": 0}

    db = get_db()
    lines = []
    errors = []

    try:
        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes), data_only=True)
        ws = wb.active
        raw_headers = [str(cell.value or "").strip().lower().replace(" ", "_") for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        has_monthly = any(h in raw_headers for h in MONTH_COLS)
        year = period[:4]  # used for monthly_amounts keys

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue
            row_dict = dict(zip(raw_headers, row))
            coa_code = str(row_dict.get("coa_code") or row_dict.get("code") or "").strip()
            if not coa_code:
                continue

            coa = await db.chart_of_accounts.find_one({"code": coa_code, "deleted_at": None})
            if not coa:
                errors.append(f"Row {i}: COA '{coa_code}' tidak ditemukan")
                continue

            category = str(row_dict.get("category") or row_dict.get("kategori") or "").strip() or guess_category(coa_code)

            if has_monthly:
                # Read monthly columns
                monthly_amounts = {}
                total = 0.0
                for mname in MONTH_COLS:
                    val = row_dict.get(mname) or 0
                    try:
                        v = float(str(val).replace(",", "").strip() or 0)
                    except (ValueError, TypeError):
                        v = 0.0
                    month_key = f"{year}-{MONTH_IDX[mname]}"
                    monthly_amounts[month_key] = v
                    total += v
                # Allow override from "total" column
                total_col = row_dict.get("total") or row_dict.get("jumlah")
                if total_col:
                    try:
                        total = float(str(total_col).replace(",", "").strip())
                    except (ValueError, TypeError):
                        pass
                lines.append({
                    "coa_id": coa["id"],
                    "coa_code": coa_code,
                    "coa_name": coa["name"],
                    "category": category,
                    "amount": round(total, 2),
                    "monthly_amounts": {k: round(v, 2) for k, v in monthly_amounts.items()},
                })
            else:
                # Simple template
                amount_raw = row_dict.get("amount") or row_dict.get("jumlah") or 0
                try:
                    amount = float(str(amount_raw).replace(",", "").strip())
                except (ValueError, TypeError):
                    errors.append(f"Row {i}: invalid amount")
                    continue
                lines.append({
                    "coa_id": coa["id"],
                    "coa_code": coa_code,
                    "coa_name": coa["name"],
                    "category": category,
                    "amount": amount,
                })
    except Exception as e:
        return {"success": False, "errors": [f"Gagal membaca Excel: {e}"], "imported": 0}

    if not lines:
        return {"success": False, "errors": errors or ["Tidak ada baris valid"], "imported": 0}

    period_type = "annual_monthly" if has_monthly else "monthly"
    doc = make_budget_doc(
        name=f"Budget Import {'Annual' if has_monthly else ''} {period}",
        period=period,
        period_type=period_type,
        scope=scope,
        outlet_id=outlet_id,
        brand_id=brand_id,
        lines=lines,
        notes=f"Imported from Excel ({len(lines)} rows{'  + monthly' if has_monthly else ''})",
        created_by=user_id,
    )
    await db.budgets.insert_one(doc)
    return {
        "success": True,
        "imported": len(lines),
        "skipped_errors": len(errors),
        "errors": errors,
        "budget_id": doc["id"],
        "period": period,
        "has_monthly": has_monthly,
    }


async def generate_template_excel(period_type: str = "simple") -> bytes:
    """Generate Excel template. period_type: simple | monthly"""
    import io as _io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Template"

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    if period_type == "monthly":
        # Monthly breakdown template
        headers = ["coa_code", "coa_name"] + MONTH_COLS + ["total", "category"]
        ws.append(headers)
        sample = [
            ["4001", "Revenue Food"] + [10000000]*12 + [120000000, "REV"],
            ["4002", "Revenue Beverage"] + [3000000]*12 + [36000000, "REV"],
            ["5001", "HPP Food"] + [4000000]*12 + [48000000, "COGS"],
            ["5401", "Gaji Karyawan"] + [8000000]*12 + [96000000, "PAYROLL"],
            ["5601", "Biaya Operasional"] + [2000000]*12 + [24000000, "OPEX"],
        ]
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 30
        for i in range(len(MONTH_COLS)):
            ws.column_dimensions[get_column_letter(i + 3)].width = 14
        ws.column_dimensions[get_column_letter(len(MONTH_COLS) + 3)].width = 16
        ws.column_dimensions[get_column_letter(len(MONTH_COLS) + 4)].width = 12
    else:
        # Simple template
        headers = ["coa_code", "coa_name", "amount", "category"]
        ws.append(headers)
        sample = [
            ["4001", "Revenue Food", 10000000, "REV"],
            ["4002", "Revenue Beverage", 3000000, "REV"],
            ["5001", "HPP Food", 4000000, "COGS"],
            ["5401", "Gaji Karyawan", 8000000, "PAYROLL"],
            ["5601", "Biaya Operasional", 2000000, "OPEX"],
        ]
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in sample:
        ws.append(row)

    # Categories sheet
    ws2 = wb.create_sheet("Categories")
    ws2.append(["Code", "Name", "Description"])
    for cat in BUDGET_CATEGORIES:
        if not cat.get("derived"):
            ws2.append([cat["code"], cat["name"], cat.get("description", "")])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
