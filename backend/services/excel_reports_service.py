"""excel_reports_service — Generate XLSX reports for Torado ERP.

Covers:
- Anomaly Feed                (no existing xlsx equivalent)
- Payroll Cycles              (no existing xlsx equivalent)
- Purchase Requests (PR)      (no existing xlsx equivalent)
- AR Write-off Report         (no existing xlsx equivalent)

NOTE: AP Aging XLSX and PO XLSX are handled by existing reports_excel_finance_service
and reports_excel_procurement_service to avoid duplication.
The route endpoints /api/finance/ap-aging/export/xlsx and
/api/procurement/pos/export/xlsx delegate to those services.
"""
from __future__ import annotations

from datetime import datetime, timezone
from typing import Optional

from core.db import get_db, serialize
from services.excel_export_service import (
    add_report_header,
    autosize_columns,
    create_workbook,
    format_date,
    freeze_header_row,
    workbook_to_bytes,
    write_data_row,
    write_table_headers,
    write_total_row,
    apply_currency_format,
    HEADER_FILL, HEADER_FONT, HEADER_ALIGNMENT, BORDER_THIN,
)
from openpyxl.styles import Font, PatternFill, Alignment

_CURRENCY_COLS_FMT = '"Rp"#,##0'


def _now_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


# ─────────────────────────────────────────────────────────────────────────────
# 1. ANOMALY FEED
# ─────────────────────────────────────────────────────────────────────────────
async def generate_anomaly_xlsx(
    *,
    type: Optional[str] = None,
    severity: Optional[str] = None,
    status: Optional[str] = None,
    outlet_id: Optional[str] = None,
    vendor_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 1000,
) -> bytes:
    db = get_db()
    q: dict = {"deleted_at": None}
    if type:
        q["type"] = type
    if severity:
        q["severity"] = severity
    if status:
        q["status"] = status
    if outlet_id:
        q["outlet_id"] = outlet_id
    if vendor_id:
        q["vendor_id"] = vendor_id
    if date_from:
        q.setdefault("scan_date", {})["$gte"] = date_from
    if date_to:
        q.setdefault("scan_date", {})["$lte"] = date_to

    docs = [serialize(d) async for d in db.anomaly_events.find(q).sort([("created_at", -1)]).limit(limit)]

    # Enrich outlet names
    outlet_map: dict = {}
    async for o in db.outlets.find({"deleted_at": None}, {"id": 1, "name": 1}):
        outlet_map[o["id"]] = o["name"]

    wb = create_workbook("Anomaly Feed")
    ws = wb.active

    subtitle_parts = []
    if date_from or date_to:
        subtitle_parts.append(f"Periode: {date_from or '—'} s/d {date_to or '—'}")
    if severity:
        subtitle_parts.append(f"Severity: {severity}")
    if status:
        subtitle_parts.append(f"Status: {status}")

    row = add_report_header(
        ws, "Anomaly Feed Report",
        subtitle=", ".join(subtitle_parts) or "Semua anomali",
        generated_at=_now_str(),
    )

    headers = ["No", "Tanggal Scan", "Tipe", "Severity", "Status", "Outlet", "Judul", "Deskripsi", "Assigned To"]
    row = write_table_headers(ws, headers, row)

    number_cols: list[int] = [1]
    for i, d in enumerate(docs, 1):
        write_data_row(ws, [
            i,
            format_date(d.get("scan_date") or d.get("created_at")),
            d.get("type", ""),
            d.get("severity", "").upper(),
            d.get("status", ""),
            outlet_map.get(d.get("outlet_id", ""), d.get("outlet_id", "")),
            d.get("title", ""),
            (d.get("description") or "")[:200],
            d.get("assigned_to", ""),
        ], row, number_cols=number_cols)

        # Color-code severity
        sev = d.get("severity", "").lower()
        color = {"critical": "FFCCCC", "high": "FFE5CC", "medium": "FFFACC", "low": "CCFFCC"}.get(sev, "FFFFFF")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row += 1

    write_total_row(ws, [f"Total: {len(docs)} anomali", "", "", "", "", "", "", "", ""], row)
    freeze_header_row(ws, header_row=row - len(docs) - 1)
    autosize_columns(ws)
    ws.column_dimensions["G"].width = 35
    ws.column_dimensions["H"].width = 45

    return workbook_to_bytes(wb)

# ─────────────────────────────────────────────────────────────────────────────
# 3. PAYROLL CYCLES
# ─────────────────────────────────────────────────────────────────────────────
async def generate_payroll_xlsx(*, period: Optional[str] = None, limit: int = 500) -> bytes:
    db = get_db()
    q: dict = {"deleted_at": None}
    if period:
        q["period"] = period

    docs = [serialize(d) async for d in db.payroll_cycles.find(q).sort([("period", -1)]).limit(limit)]

    outlet_map: dict = {}
    async for o in db.outlets.find({"deleted_at": None}, {"id": 1, "name": 1}):
        outlet_map[o["id"]] = o["name"]

    wb = create_workbook("Payroll")
    ws = wb.active
    row = add_report_header(
        ws, "Payroll Cycles Report",
        subtitle=f"Periode: {period or 'Semua'}",
        generated_at=_now_str(),
    )

    headers = ["No", "Periode", "Outlet", "Total Karyawan", "Gaji Pokok", "BPJS TK (TK)", "BPJS TK (Emp)", "BPJS Kes (TK)", "BPJS Kes (Emp)", "PPh21", "Total Nett", "Status"]
    row = write_table_headers(ws, headers, row)
    data_start = row
    number_cols = [4, 5, 6, 7, 8, 9, 10, 11]

    total_gross, total_nett, total_emp = 0.0, 0.0, 0
    for i, d in enumerate(docs, 1):
        gross = float(d.get("total_gross", 0) or 0)
        nett = float(d.get("total_nett", 0) or 0)
        emp = int(d.get("employee_count", 0) or 0)
        total_gross += gross
        total_nett += nett
        total_emp += emp
        write_data_row(ws, [
            i,
            d.get("period", ""),
            outlet_map.get(d.get("outlet_id", ""), d.get("outlet_id", "")),
            emp,
            gross,
            float(d.get("bpjs_tk_employer", 0) or 0),
            float(d.get("bpjs_tk_employee", 0) or 0),
            float(d.get("bpjs_kes_employer", 0) or 0),
            float(d.get("bpjs_kes_employee", 0) or 0),
            float(d.get("pph21_total", 0) or 0),
            nett,
            d.get("status", ""),
        ], row, number_cols=number_cols)
        row += 1

    write_total_row(ws, ["TOTAL", "", "", total_emp, total_gross, "", "", "", "", "", total_nett, ""], row, number_cols=number_cols)
    for col in number_cols:
        apply_currency_format(ws, col, data_start, row)
    freeze_header_row(ws, header_row=data_start - 1)
    autosize_columns(ws)
    return workbook_to_bytes(wb)


# ─────────────────────────────────────────────────────────────────────────────
# 4. PURCHASE REQUESTS (PR)
# ─────────────────────────────────────────────────────────────────────────────
async def generate_pr_xlsx(
    *,
    status: Optional[str] = None,
    outlet_id: Optional[str] = None,
    source: Optional[str] = None,
    limit: int = 1000,
) -> bytes:
    db = get_db()
    q: dict = {"deleted_at": None}
    if status:
        q["status"] = status
    if outlet_id:
        q["outlet_id"] = outlet_id
    if source:
        q["source"] = source

    docs = [serialize(d) async for d in db.purchase_requests.find(q).sort([("created_at", -1)]).limit(limit)]

    outlet_map: dict = {}
    async for o in db.outlets.find({"deleted_at": None}, {"id": 1, "name": 1}):
        outlet_map[o["id"]] = o["name"]

    wb = create_workbook("Purchase Requests")
    ws = wb.active

    parts = []
    if status:
        parts.append(f"Status: {status}")
    if source:
        parts.append(f"Source: {source}")

    row = add_report_header(
        ws, "Purchase Request Report",
        subtitle=", ".join(parts) or "Semua PR",
        generated_at=_now_str(),
    )

    headers = ["No", "Doc No", "Tanggal", "Outlet", "Source", "Jumlah Lines", "Status", "Catatan"]
    row = write_table_headers(ws, headers, row)
    number_cols = [1, 6]

    for i, d in enumerate(docs, 1):
        write_data_row(ws, [
            i,
            d.get("doc_no", d.get("id", "")[:8]),
            format_date(d.get("request_date")),
            outlet_map.get(d.get("outlet_id", ""), d.get("outlet_id", "")),
            d.get("source", "manual"),
            len(d.get("lines") or []),
            d.get("status", ""),
            (d.get("notes") or "")[:100],
        ], row, number_cols=number_cols)
        row += 1

    write_total_row(ws, [f"Total: {len(docs)} PR", "", "", "", "", "", "", ""], row)
    freeze_header_row(ws, header_row=row - len(docs) - 1)
    autosize_columns(ws)
    return workbook_to_bytes(wb)


# ─────────────────────────────────────────────────────────────────────────────
# 5. AR WRITE-OFF REPORT
# ─────────────────────────────────────────────────────────────────────────────
async def generate_ar_writeoff_xlsx(
    *,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 1000,
) -> bytes:
    """Export AR write-off history as XLSX."""
    db = get_db()
    q: dict = {"status": "written_off", "deleted_at": None}
    if date_from:
        q.setdefault("write_off_date", {})["$gte"] = date_from
    if date_to:
        q.setdefault("write_off_date", {})["$lte"] = date_to

    docs = [serialize(d) async for d in db.ar_invoices.find(q).sort([("write_off_date", -1)]).limit(limit)]

    customer_map: dict = {}
    async for c in db.ar_customers.find({"deleted_at": None}, {"id": 1, "name": 1}):
        customer_map[c["id"]] = c["name"]

    wb = create_workbook("AR Write-off")
    ws = wb.active

    parts = []
    if date_from or date_to:
        parts.append(f"Periode: {date_from or '—'} s/d {date_to or '—'}")

    row = add_report_header(
        ws, "AR Write-off Report",
        subtitle=", ".join(parts) or "Semua periode",
        generated_at=_now_str(),
    )

    headers = ["No", "No Invoice", "Tgl Invoice", "Tgl Write-off", "Customer", "Nominal", "Alasan", "JE ID"]
    row = write_table_headers(ws, headers, row)
    data_start = row
    number_cols = [1, 6]

    total = 0.0
    for i, d in enumerate(docs, 1):
        amt = float(d.get("total_amount", 0) or d.get("amount", 0) or 0)
        total += amt
        write_data_row(ws, [
            i,
            d.get("invoice_no", ""),
            format_date(d.get("invoice_date")),
            format_date(d.get("write_off_date")),
            customer_map.get(d.get("customer_id", ""), d.get("customer_id", "")),
            amt,
            (d.get("write_off_reason") or "")[:100],
            d.get("write_off_je_id", ""),
        ], row, number_cols=number_cols)
        row += 1

    write_total_row(ws, [f"Total: {len(docs)} invoice", "", "", "", "", total, "", ""], row, number_cols=number_cols)
    apply_currency_format(ws, 6, data_start, row)
    freeze_header_row(ws, header_row=data_start - 1)
    autosize_columns(ws)
    ws.column_dimensions["G"].width = 40
    return workbook_to_bytes(wb)

