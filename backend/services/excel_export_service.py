"""Excel Export Service — Universal utilities for generating styled .xlsx reports.

Provides:
- Workbook creation with Torado branding
- Consistent styling (headers, data cells, number formats)
- Utility functions: autosize columns, freeze panes, add totals, conditional formatting
- Chart helpers (bar, line)
"""
from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

logger = logging.getLogger("aurora.excel_export")


# ============================================================
# STYLING CONSTANTS (Torado Brand Colors)
# ============================================================
TORADO_GOLD = "C9A876"
TORADO_ESPRESSO = "1C1510"
TORADO_DARK_BG = "14100B"
TORADO_LIGHT_TEXT = "F0EAE0"

HEADER_FILL = PatternFill(start_color=TORADO_ESPRESSO, end_color=TORADO_ESPRESSO, fill_type="solid")
HEADER_FONT = Font(name="Manrope", size=11, bold=True, color=TORADO_LIGHT_TEXT)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(start_color=TORADO_DARK_BG, end_color=TORADO_DARK_BG, fill_type="solid")
SUBHEADER_FONT = Font(name="Manrope", size=10, bold=True, color=TORADO_LIGHT_TEXT)

DATA_FONT = Font(name="Manrope", size=10, color="000000")
DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center")

NUMBER_FONT = Font(name="IBM Plex Mono", size=10, color="000000")
NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

TOTAL_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
TOTAL_FONT = Font(name="Manrope", size=10, bold=True, color="000000")

BORDER_THIN = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


# ============================================================
# CORE WORKBOOK CREATION
# ============================================================
def create_workbook(title: str = "Report") -> Workbook:
    """Create a new workbook with default styling."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit
    return wb


def add_report_header(
    ws,
    title: str,
    subtitle: Optional[str] = None,
    generated_at: Optional[str] = None,
    start_row: int = 1,
) -> int:
    """Add branded report header block.
    
    Returns: next available row number after header.
    """
    # Title
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    cell = ws.cell(row=start_row, column=1)
    cell.value = title
    cell.font = Font(name="Cormorant Garamond", size=16, bold=True, color=TORADO_ESPRESSO)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    
    current_row = start_row + 1
    
    # Subtitle
    if subtitle:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        cell = ws.cell(row=current_row, column=1)
        cell.value = subtitle
        cell.font = Font(name="Manrope", size=11, color="666666")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1
    
    # Generated timestamp
    if generated_at is None:
        generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
    cell = ws.cell(row=current_row, column=1)
    cell.value = f"Generated: {generated_at}"
    cell.font = Font(name="Manrope", size=9, italic=True, color="999999")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    current_row += 1
    
    # Blank row separator
    current_row += 1
    
    return current_row


def write_table_headers(
    ws,
    headers: list[str],
    start_row: int,
    start_col: int = 1,
) -> int:
    """Write table headers with styling.
    
    Returns: next row number (data row).
    """
    for idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row, column=idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = BORDER_THIN
    
    return start_row + 1


def write_data_row(
    ws,
    row_data: list[Any],
    row_num: int,
    start_col: int = 1,
    number_cols: Optional[list[int]] = None,
) -> None:
    """Write a single data row with appropriate styling.
    
    Args:
        number_cols: List of column indices (1-based) that contain numbers (right-aligned).
    """
    if number_cols is None:
        number_cols = []
    
    for idx, value in enumerate(row_data, start=start_col):
        cell = ws.cell(row=row_num, column=idx)
        cell.value = value
        
        if idx in number_cols:
            cell.font = NUMBER_FONT
            cell.alignment = NUMBER_ALIGNMENT
            # Apply number format if value is numeric
            if isinstance(value, (int, float)):
                if isinstance(value, int) and abs(value) < 1_000_000:
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "#,##0.00"
        else:
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
        
        cell.border = BORDER_THIN


def write_total_row(
    ws,
    row_data: list[Any],
    row_num: int,
    start_col: int = 1,
    number_cols: Optional[list[int]] = None,
) -> None:
    """Write a total/summary row with bold styling."""
    if number_cols is None:
        number_cols = []
    
    for idx, value in enumerate(row_data, start=start_col):
        cell = ws.cell(row=row_num, column=idx)
        cell.value = value
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = BORDER_THIN
        
        if idx in number_cols:
            cell.alignment = NUMBER_ALIGNMENT
            if isinstance(value, (int, float)):
                if isinstance(value, int) and abs(value) < 1_000_000:
                    cell.number_format = "#,##0"
                else:
                    cell.number_format = "#,##0.00"
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")


def autosize_columns(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        adjusted_width = min(max(length + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width


def freeze_header_row(ws, header_row: int = 1) -> None:
    """Freeze panes at header row."""
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def apply_currency_format(ws, col_num: int, start_row: int, end_row: int, currency: str = "Rp") -> None:
    """Apply currency number format to a column range."""
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col_num)
        if isinstance(cell.value, (int, float)):
            if currency == "Rp":
                cell.number_format = '"Rp"#,##0'
            else:
                cell.number_format = f'"{currency}"#,##0.00'


def add_bar_chart(
    ws,
    title: str,
    data_range: str,
    categories_range: str,
    position: str = "H2",
) -> None:
    """Add a simple bar chart to the worksheet.
    
    Args:
        data_range: e.g., "B2:B10"
        categories_range: e.g., "A2:A10"
        position: e.g., "H2" (top-left corner of chart)
    """
    chart = BarChart()
    chart.title = title
    chart.style = 10
    chart.height = 10  # default 7.5
    chart.width = 20   # default 15
    
    # Prefix sheet name to range for openpyxl Reference compatibility
    sheet_name = ws.title
    data_qualified = data_range if "!" in data_range else f"'{sheet_name}'!{data_range}"
    cats_qualified = categories_range if "!" in categories_range else f"'{sheet_name}'!{categories_range}"
    
    data = Reference(ws, range_string=data_qualified)
    cats = Reference(ws, range_string=cats_qualified)
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(cats)
    
    ws.add_chart(chart, position)


# ============================================================
# WORKBOOK TO BYTES (for download response)
# ============================================================
def workbook_to_bytes(wb: Workbook) -> bytes:
    """Convert workbook to bytes for HTTP response."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ============================================================
# HELPER: Format date for display
# ============================================================
def format_date(date_str: Optional[str]) -> str:
    """Format ISO date to display format (DD/MM/YYYY)."""
    if not date_str:
        return ""
    try:
        dt = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y")
    except Exception:  # noqa: BLE001
        return date_str


def format_datetime(dt_str: Optional[str]) -> str:
    """Format ISO datetime to display format (DD/MM/YYYY HH:MM)."""
    if not dt_str:
        return ""
    try:
        dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:  # noqa: BLE001
        return dt_str


def format_rupiah(amount: float) -> str:
    """Format number as Rupiah string."""
    return f"Rp {amount:,.0f}".replace(",", ".")
