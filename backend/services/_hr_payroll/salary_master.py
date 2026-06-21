"""Salary Master CRUD and import."""
from __future__ import annotations

import csv
import uuid
from typing import Optional

from core.db import get_db, serialize
from core.exceptions import NotFoundError
from services.hr_constants import _calc_bpjs, _calc_pph21, _now, PTKP_BY_STATUS

STANDARD_COMPONENTS = [
    {"code": "TUNJ_JABATAN",   "name": "Tunjangan Jabatan",    "amount": 0.0},
    {"code": "TUNJ_MAKAN",     "name": "Tunjangan Makan",      "amount": 0.0},
    {"code": "TUNJ_TRANSPORT", "name": "Tunjangan Transport",  "amount": 0.0},
    {"code": "TUNJ_KESEHATAN", "name": "Tunjangan Kesehatan",  "amount": 0.0},
]
PTKP_OPTIONS = list(PTKP_BY_STATUS.keys())


async def get_salary_master(employee_id: str) -> dict | None:
    db = get_db()
    emp = await db.employees.find_one({"id": employee_id, "deleted_at": None})
    if not emp:
        raise NotFoundError("Karyawan tidak ditemukan")
    sm = await db.salary_masters.find_one({"employee_id": employee_id, "deleted_at": None})
    if not sm:
        basic = float(emp.get("basic_salary", 0) or 0)
        return {
            "employee_id": employee_id, "employee_name": emp.get("full_name"),
            "outlet_id": emp.get("outlet_id"), "basic_salary": basic,
            "components": STANDARD_COMPONENTS.copy(), "bpjs_enrolled": True,
            "ptkp_status": "TK/0", "npwp": emp.get("npwp", ""), "notes": "", "is_default": True,
        }
    return serialize({**sm, "employee_name": emp.get("full_name"), "outlet_id": emp.get("outlet_id"), "is_default": False})


async def set_salary_master(employee_id: str, payload: dict, *, user: dict) -> dict:
    db = get_db()
    emp = await db.employees.find_one({"id": employee_id, "deleted_at": None})
    if not emp:
        raise NotFoundError("Karyawan tidak ditemukan")
    now = _now()
    cleaned = [{"code": c.get("code", "MISC"), "name": c.get("name", ""), "amount": float(c.get("amount", 0) or 0)} for c in payload.get("components", [])]
    doc = {
        "basic_salary": float(payload.get("basic_salary", emp.get("basic_salary", 0)) or 0),
        "components": cleaned, "bpjs_enrolled": bool(payload.get("bpjs_enrolled", True)),
        "ptkp_status": payload.get("ptkp_status", "TK/0"),
        "npwp": (payload.get("npwp") or emp.get("npwp", "")).strip(),
        "notes": payload.get("notes", ""), "updated_at": now, "updated_by": user["id"],
    }
    existing = await db.salary_masters.find_one({"employee_id": employee_id, "deleted_at": None})
    if existing:
        await db.salary_masters.update_one({"id": existing["id"]}, {"$set": doc})
        updated = await db.salary_masters.find_one({"id": existing["id"]})
        return serialize(updated)
    else:
        doc["id"] = str(uuid.uuid4())
        doc["employee_id"] = employee_id
        doc["created_at"] = now
        doc["created_by"] = user["id"]
        doc["deleted_at"] = None
        await db.salary_masters.insert_one(doc)
        return serialize(doc)


async def list_salary_masters(outlet_id: str | None = None, per_page: int = 100) -> list:
    db = get_db()
    emp_filter: dict = {"deleted_at": None, "status": "active"}
    if outlet_id:
        emp_filter["outlet_id"] = outlet_id
    emps = await db.employees.find(emp_filter).to_list(per_page)
    emp_ids_page = [e["id"] for e in emps]
    sms_page = await db.salary_masters.find({"employee_id": {"$in": emp_ids_page}, "deleted_at": None}).to_list(len(emp_ids_page) + 1) if emp_ids_page else []
    sm_page_map = {s["employee_id"]: s for s in sms_page}
    result = []
    for emp in emps:
        sm = sm_page_map.get(emp["id"])
        basic = float(emp.get("basic_salary", 0) or 0)
        if sm:
            basic = float(sm.get("basic_salary", basic) or basic)
        allowances_total = sum(float(c.get("amount", 0) or 0) for c in (sm or {}).get("components", []))
        bpjs = _calc_bpjs(basic + allowances_total, (sm or {}).get("bpjs_enrolled", True))
        pph21_detail = _calc_pph21(basic + allowances_total, (sm or {}).get("ptkp_status", "TK/0"))
        result.append({
            "employee_id": emp["id"], "employee_name": emp.get("full_name"),
            "employee_code": emp.get("employee_id"), "outlet_id": emp.get("outlet_id"),
            "position": emp.get("position"), "basic_salary": basic,
            "allowances_total": allowances_total, "total_fixed_pay": round(basic + allowances_total, 2),
            "bpjs_enrolled": (sm or {}).get("bpjs_enrolled", True),
            "bpjs_employee": bpjs["employee"], "bpjs_employer": bpjs["employer"],
            "ptkp_status": (sm or {}).get("ptkp_status", "TK/0"),
            "pph21_monthly": pph21_detail["monthly_tax"],
            "npwp": (sm or {}).get("npwp") or emp.get("npwp", ""),
            "has_master": sm is not None,
        })
    return result


async def import_salary_excel(file_bytes: bytes, *, user: dict) -> dict:
    """Import salary master from Excel (.xlsx) or CSV."""
    import io as _io
    db = get_db()
    errors = []
    rows_in = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(_io.BytesIO(file_bytes))
        ws = wb.active
        headers = [str(cell.value or "").strip().lower().replace(" ", "_") for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if all(v is None for v in row):
                continue
            rows_in.append({"_row": i, **dict(zip(headers, row))})
    except Exception:
        try:
            reader = csv.DictReader(_io.StringIO(file_bytes.decode("utf-8-sig")))
            for i, row in enumerate(reader, start=2):
                rows_in.append({"_row": i, **{k.strip().lower().replace(" ", "_"): v for k, v in row.items()}})
        except Exception as e:
            return {"imported": 0, "updated": 0, "errors": [f"Cannot parse file: {e}"], "preview": []}

    imported = updated = 0
    preview = []
    for row in rows_in:
        rn = row.get("_row", "?")
        emp = None
        for key in ("employee_code", "employee_id", "kode_karyawan", "nik"):
            val = str(row.get(key) or "").strip()
            if val:
                emp = await db.employees.find_one({"$or": [{"employee_id": val}, {"id": val}, {"npwp": val}], "deleted_at": None})
                if emp:
                    break
        if not emp:
            name_val = str(row.get("full_name") or row.get("nama") or "").strip()
            if name_val:
                emp = await db.employees.find_one({"full_name": {"$regex": name_val, "$options": "i"}, "deleted_at": None})
        if not emp:
            errors.append(f"Row {rn}: Employee not found")
            continue
        def _float(key, default=0.0):
            v = row.get(key)
            try:
                return float(str(v).replace(",", "").strip()) if v not in (None, "") else default
            except Exception:
                return default
        basic = _float("basic_salary") or _float("gaji_pokok") or float(emp.get("basic_salary", 0) or 0)
        components = [{"code": code, "name": name, "amount": _float(col)} for code, name, col in [
            ("TUNJ_JABATAN", "Tunjangan Jabatan", "tunjangan_jabatan"),
            ("TUNJ_MAKAN", "Tunjangan Makan", "tunjangan_makan"),
            ("TUNJ_TRANSPORT", "Tunjangan Transport", "tunjangan_transport"),
            ("TUNJ_KESEHATAN", "Tunjangan Kesehatan", "tunjangan_kesehatan"),
        ]]
        bpjs_val = str(row.get("bpjs_enrolled", "true") or "true").lower().strip()
        bpjs_enrolled = bpjs_val not in ("false", "0", "no", "tidak")
        ptkp_status = str(row.get("ptkp_status") or "TK/0").strip().upper()
        if ptkp_status not in PTKP_BY_STATUS:
            ptkp_status = "TK/0"
        npwp = str(row.get("npwp") or emp.get("npwp", "") or "").strip()
        existing = await db.salary_masters.find_one({"employee_id": emp["id"], "deleted_at": None})
        now = _now()
        if existing:
            await db.salary_masters.update_one({"id": existing["id"]}, {"$set": {"basic_salary": basic, "components": components, "bpjs_enrolled": bpjs_enrolled, "ptkp_status": ptkp_status, "npwp": npwp, "updated_at": now, "updated_by": user["id"]}})
            updated += 1
        else:
            await db.salary_masters.insert_one({"id": str(uuid.uuid4()), "employee_id": emp["id"], "basic_salary": basic, "components": components, "bpjs_enrolled": bpjs_enrolled, "ptkp_status": ptkp_status, "npwp": npwp, "notes": "", "created_at": now, "created_by": user["id"], "updated_at": now, "updated_by": user["id"], "deleted_at": None})
            imported += 1
        preview.append({"employee_name": emp.get("full_name"), "employee_id": emp["id"], "basic_salary": basic, "allowances_total": sum(c["amount"] for c in components), "ptkp_status": ptkp_status, "bpjs_enrolled": bpjs_enrolled})
    return {"imported": imported, "updated": updated, "errors": errors, "preview": preview}
